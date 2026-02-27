# ==============================================================================
# ENHANCED OPTION BUYER SCANNER v4.0 - MERGED VERSION WITH OPTION CHAIN INTEGRATION
# Integrates TrueData for TA and localhost API for option chain data
# Parallel fetching for efficiency
# Adds option chain metrics to scoring for option buyers focus
# ==============================================================================

import os
import logging
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import pytz
import time
import threading
from collections import defaultdict
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json
import requests
import random

from tqdm import tqdm
from truedata_ws.websocket.TD import TD  # Updated import for current library

# Enhanced table formatting libraries
try:
    from rich.console import Console
    from rich.table import Table
    from rich.text import Text
    from rich import box
    from rich.panel import Panel
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False
    print("Install rich: pip install rich")

try:
    from colorama import init, Fore, Back, Style
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False
    print("Install colorama: pip install colorama")

try:
    from great_tables import GT, md, html, style, loc
    from great_tables.data import sp500
    GREAT_TABLES_AVAILABLE = True
except ImportError:
    GREAT_TABLES_AVAILABLE = False
    print("Install great-tables: pip install great-tables")

try:
    from tabulate import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False

# Initialize console for rich output
if RICH_AVAILABLE:
    console = Console()

# Create a simple logger replacement
class Logger:
    def info(self, msg): print(f"[INFO] {msg}")
    def error(self, msg): print(f"[ERROR] {msg}")
    def warning(self, msg): print(f"[WARNING] {msg}")
    def exception(self, msg): print(f"[EXCEPTION] {msg}")

logger = Logger()

# ======== Enhanced Configuration ========
class Config:
    TDUSERNAME = os.getenv("TRUEDATA_USER", "tdwsp751")  # Replace with real credentials
    TDPASSWORD = os.getenv("TRUEDATA_PASS", "raj@751")  # Replace with real credentials

    MARKET_START = "09:15"  # IST
    FIRST_RUN_AT = "09:20"  # IST; First scan after 09:15-09:20 candle
    FIRST_SCAN_DELAY = 15   # Wait 15 seconds after 09:20 for settlement
    MARKET_END   = "15:30"  # IST
    SETTLE_DELAY_SECONDS = 15  # wait after bar close for data settlement
    MAX_WORKERS = int(os.getenv("MAX_WORKERS", "64"))
    TD_HIST_SESSIONS = int(os.getenv("TD_HIST_SESSIONS", "5"))
    SHARES_FILE = os.getenv("SHARES_FILE", "shares.txt")
    BENCHMARK_INDEX = "NIFTY 50"

    # Option Chain API Config (from CE_PE_1.py)
    OPTION_API_TMPL = "http://localhost:3000/api/equity/options/{symbol}"
    OPTION_TIMEOUT = 20
    MIN_TOTAL_OI = 2000
    MIN_TOTAL_VOL = 200
    PCR_TOL = 0.03
    EPS = 1e-6

    # --- Backtesting Configuration ---
    BACKTEST_INTERVAL_MINUTES = 5
    BACKTEST_START_DELAY = 5
    BACKTEST_TOP_DISPLAY = 15  # Updated to 15 as per task
    
    # --- Indicator Group Weights ---
    GROUP_WEIGHTS = {
        "Trend": 2.5, "Momentum": 2.0, "Volume": 2.2, "Volatility": 1.8, "OI": 2.5,
        "Option": 3.0  # New group for option chain data, high weight for option buyers
    }

    # --- Individual Indicator Weights within Groups ---
    INDICATOR_WEIGHTS = {
        "MA_Slope": 2.0, "ADX": 1.8, "VWAP": 1.7, "EMA": 1.7, "MACD_Trend": 1.5,
        "RSI": 2.0, "Stochastic": 1.2, "CCI": 1.2, "ROC": 1.1, "WilliamsR": 1.0,
        "VolumeSurge": 2.5, "OBV": 1.8, "CMF": 1.8, "RelVol": 1.5,
        "VolatilityExpansion": 2.5, "Bollinger": 1.3,
        "OptionBuyerMomentum": 2.8, "OIChange": 2.5, "VolumeOISync": 2.2,
        "OptionSentiment": 2.5  # New for option chain remark-based sentiment
    }

    # --- Scoring & Signal Thresholds ---
    SCORE_THRESHOLD_MIN = 10.0
    SIGNAL_THRESHOLDS = {
        'Very Strong Buy': 55.0, 'Strong Buy': 30.0, 'Buy Signal': 15.0,
        'Very Strong Sell': -55.0, 'Strong Sell': -30.0, 'Sell Signal': -15.0,
    }
    
    # --- Market Regime Multipliers ---
    REGIME_MULTIPLIERS = {
        'bullish_in_bull_market': 1.15, 'bearish_in_bear_market': 1.15,
        'bullish_in_bear_market': 0.8, 'bearish_in_bull_market': 0.8,
    }

# Constants
IST = pytz.timezone("Asia/Kolkata")
BAR_SIZE_MAP = {5: "5 min", 15: "15 min", 30: "30 min", 60: "60 min", 1440: "1 day"}
DURATION_MAP = {5: "30 D", 15: "30 D", 30: "60 D", 60: "120 D", 1440: "365 D"}
TIMEFRAME_WEIGHTS = {15: 3.0, 5: 2.5, 30: 2.0, 60: 1.5, 1440: 1.0}

# Silence noisy loggers
for noisy in ("truedata", "truedata_ws", "websocket", "urllib3"):
    logging.getLogger(noisy).setLevel(logging.CRITICAL)

# TrueData sessions pool (updated to TD class)
td_pool = [TD(Config.TDUSERNAME, Config.TDPASSWORD) for _ in range(Config.TD_HIST_SESSIONS)]

# State management
previous_scan_results = {}
previous_oi_data = {}
previous_volume_data = {}
intraday_volume_data = {}  # Track 5-minute volume changes
intraday_oi_data = {}      # Track 5-minute OI changes
scan_count = 0
backtest_stock_history = {}
current_scan_data = {}

# Color definitions
class Colors:
    HEADER = '\033[95m'; BLUE = '\033[94m'; CYAN = '\033[96m'
    GREEN = '\033[92m'; YELLOW = '\033[93m'; RED = '\033[91m'
    BOLD = '\033[1m'; UNDERLINE = '\033[4m'; END = '\033[0m'
    MAGENTA = '\033[35m'; ORANGE = '\033[33m'

def print_colored(text, color):
    if COLORAMA_AVAILABLE:
        color_map = {
            Colors.HEADER: Fore.MAGENTA + Style.BRIGHT,
            Colors.BLUE: Fore.BLUE + Style.BRIGHT,
            Colors.CYAN: Fore.CYAN + Style.BRIGHT,
            Colors.GREEN: Fore.GREEN + Style.BRIGHT,
            Colors.YELLOW: Fore.YELLOW + Style.BRIGHT,
            Colors.RED: Fore.RED + Style.BRIGHT,
            Colors.BOLD: Style.BRIGHT,
            Colors.MAGENTA: Fore.MAGENTA + Style.BRIGHT,
            Colors.ORANGE: Fore.YELLOW + Style.BRIGHT,
        }
        print(color_map.get(color, '') + text)
    else:
        print(f"{color}{text}{Colors.END}")

# ========== OPTION CHAIN FUNCTIONS FROM CE_PE_1.py ==========
def safe_div(a, b):
    if b is None or abs(b) < Config.EPS:
        return float('inf') if a > 0 else 0.0
    return a / b

def parse_expiry(s):
    try:
        return datetime.strptime(s, "%d-%b-%Y")
    except (ValueError, TypeError):
        return None

def choose_current_expiry(records):
    exps = records.get("expiryDates") or []
    exps_parsed = [(e, parse_expiry(e)) for e in exps]
    now = datetime.now()
    future = [e for e in exps_parsed if e[1] and e[1] >= now]
    if future:
        return min(future, key=lambda x: x[1])[0]
    past = [e for e in exps_parsed if e[1]]
    if past:
        return max(past, key=lambda x: x[1])[0]
    return None

def fetch_symbol_metrics(symbol):
    url = Config.OPTION_API_TMPL.format(symbol=symbol)
    try:
        r = requests.get(url, timeout=Config.OPTION_TIMEOUT)
        r.raise_for_status()
        obj = r.json()

        recs = obj.get("records", {})
        curr_exp = choose_current_expiry(recs)
        if not curr_exp:
            return {"Remark": "No valid expiry"}

        rows = [row for row in recs.get("data", []) if row.get("expiryDate") == curr_exp]
        if not rows:
            return {"Remark": "No rows for current expiry"}

        underlying = next((val for row in rows for val in [row.get("CE", {}).get("underlyingValue"), row.get("PE", {}).get("underlyingValue")] if isinstance(val, (int, float))), None)
        if underlying is None:
            return {"Remark": "Underlying price not found"}

        ce_oi_sum, pe_oi_sum = 0, 0
        ce_vol_sum, pe_vol_sum = 0, 0
        ce_oi_wsum, pe_oi_wsum = 0.0, 0.0
        ce_oi_w, pe_oi_w = 0.0, 0.0
        ce_iv_wsum, pe_iv_wsum = 0.0, 0.0
        ce_iv_w, pe_iv_w = 0.0, 0.0

        for row in rows:
            ce = row.get("CE") or {}
            pe = row.get("PE") or {}
            ce_oi = ce.get("openInterest") or 0
            pe_oi = pe.get("openInterest") or 0
            ce_vol = ce.get("totalTradedVolume") or 0
            pe_vol = pe.get("totalTradedVolume") or 0

            ce_oi_sum += ce_oi
            pe_oi_sum += pe_oi
            ce_vol_sum += ce_vol
            pe_vol_sum += pe_vol

            if isinstance(ce.get("pchangeinOpenInterest"), (int, float)) and ce_oi > 0:
                ce_oi_wsum += ce.get("pchangeinOpenInterest") * ce_oi
                ce_oi_w += ce_oi
            if isinstance(pe.get("pchangeinOpenInterest"), (int, float)) and pe_oi > 0:
                pe_oi_wsum += pe.get("pchangeinOpenInterest") * pe_oi
                pe_oi_w += pe_oi

            ce_iv = ce.get("impliedVolatility") or 0
            pe_iv = pe.get("impliedVolatility") or 0
            if ce_iv > 0 and ce_oi > 0:
                ce_iv_wsum += ce_iv * ce_oi
                ce_iv_w += ce_oi
            if pe_iv > 0 and pe_oi > 0:
                pe_iv_wsum += pe_iv * pe_oi
                pe_iv_w += pe_oi

        total_oi = ce_oi_sum + pe_oi_sum
        total_vol = ce_vol_sum + pe_vol_sum
        pcr = safe_div(pe_oi_sum, ce_oi_sum)

        ce_oi_chg_pct = safe_div(ce_oi_wsum, ce_oi_w)
        pe_oi_chg_pct = safe_div(pe_oi_wsum, pe_oi_w)
        blended_oi_chg = safe_div((ce_oi_chg_pct * ce_oi_sum) + (pe_oi_chg_pct * pe_oi_sum), total_oi)

        avg_ce_iv = safe_div(ce_iv_wsum, ce_iv_w)
        avg_pe_iv = safe_div(pe_iv_wsum, pe_iv_w)
        avg_iv = safe_div((avg_ce_iv * ce_oi_sum) + (avg_pe_iv * pe_oi_sum), total_oi) * 100

        vol_oi_ratio = safe_div(total_vol, total_oi)

        atm_strike_row = min(rows, key=lambda r: abs(r.get("strikePrice", float('inf')) - underlying))
        atm_ce = atm_strike_row.get("CE", {})
        atm_pe = atm_strike_row.get("PE", {})
        atm_pcr = safe_div(atm_pe.get("openInterest", 0), atm_ce.get("openInterest", 0))
        atm_ce_vol = atm_ce.get("totalTradedVolume", 0)
        atm_pe_vol = atm_pe.get("totalTradedVolume", 0)
        atm_vol_dom = "CALLS" if atm_ce_vol > atm_pe_vol else ("PUTS" if atm_pe_vol > atm_ce_vol else "NEUTRAL")
        atm_signal = f"PCR:{atm_pcr:.2f}|VOL:{atm_vol_dom}"

        def classify(pcr, ce_oi, pe_oi, ce_vol, pe_vol):
            is_low_liq = (ce_oi + pe_oi < Config.MIN_TOTAL_OI) or (ce_vol + pe_vol < Config.MIN_TOTAL_VOL)
            if abs(pcr - 1.0) <= Config.PCR_TOL: return "Neutral"
            ce_oi_dom, pe_oi_dom = ce_oi > pe_oi, pe_oi > ce_oi
            ce_vol_dom, pe_vol_dom = ce_vol >= pe_vol, pe_vol >= ce_vol
            if pcr < 0.8 and ce_oi_dom and ce_vol_dom: return "Strong Bullish" if not is_low_liq else "Mild Bullish"
            if pcr > 1.2 and pe_oi_dom and pe_vol_dom: return "Strong Bearish" if not is_low_liq else "Mild Bearish"
            if pcr < 1.0 and (ce_oi_dom or ce_vol_dom): return "Mild Bullish"
            if pcr > 1.0 and (pe_oi_dom or pe_vol_dom): return "Mild Bearish"
            return "Neutral"

        remark = classify(pcr, ce_oi_sum, pe_oi_sum, ce_vol_sum, pe_vol_sum)

        return {
            "Price": underlying, "Volume": total_vol, "OI": total_oi,
            "OI Chg %": blended_oi_chg, "PCR": pcr, "Avg IV %": avg_iv, "V/OI Ratio": vol_oi_ratio,
            "ATM Signal": atm_signal, "Remark": remark, "Expiry": curr_exp
        }
    except Exception as e:
        logger.error(f"Option chain fetch error for {symbol}: {e}")
        return {"Remark": f"Error: {str(e)}"}

# ========== FIXED GREAT TABLES INTEGRATION ==========
def create_great_table_fixed(data, title, new_stocks=None, show_time=None):
    if not data or not GREAT_TABLES_AVAILABLE:
        create_rich_enhanced_table(data, title, new_stocks, show_time)
        return
    try:
        df_data = []
        for item in data:
            row = {
                'Stock': item['symbol'],
                'Signal': item['signal'],
                'Score': round(item['score'], 2),
                'Trend': round(item['sub_scores'].get('Trend', 0), 2),
                'Momentum': round(item['sub_scores'].get('Momentum', 0), 2),
                'Volume': round(item['sub_scores'].get('Volume', 0), 2),
                'OI': round(item['sub_scores'].get('OI', 0), 2),
                'Option': round(item['sub_scores'].get('Option', 0), 2),  # New
                'Curr_Vol': item.get('current_volume', 'N/A'),
                'Curr_OI': item.get('current_oi', 'N/A'),
                'Vol_Change': item.get('volume_change_pct', 0),
                'OI_Change': item.get('oi_change_pct', 0),
                'PCR': round(item.get('pcr', 0), 2),  # New
                'OI_Chg_Pct': round(item.get('oi_change_pct', 0), 2),  # New
                'Avg_IV': round(item.get('avg_iv', 0), 2),  # New
                'V_OI': round(item.get('v_oi_ratio', 0), 2),  # New
                'Option_Remark': item.get('option_remark', 'Neutral'),  # New
                'Flow': item.get('flow', 'Unknown'),
                'Action': item.get('action', 'Consider'),
                'Is_New': 1 if (new_stocks and item['symbol'] in new_stocks) else 0
            }
            df_data.append(row)
        df = pd.DataFrame(df_data)
        gt_table = (
            GT(df)
            .tab_header(
                title=title,
                subtitle=f"Scan Time: {show_time}" if show_time else "Live Scanner Results"
            )
            .tab_spanner(label="Signal Analysis", columns=["Stock", "Signal", "Score"])
            .tab_spanner(label="Technical Indicators", columns=["Trend", "Momentum", "Volume", "OI", "Option"])
            .tab_spanner(label="Current Data", columns=["Curr_Vol", "Curr_OI", "Vol_Change", "OI_Change"])
            .tab_spanner(label="Option Chain", columns=["PCR", "OI_Chg_Pct", "Avg_IV", "V_OI", "Option_Remark"])  # New
            .tab_spanner(label="Analysis", columns=["Flow", "Action"])
            .fmt_number(columns=["Score", "Trend", "Momentum", "Volume", "OI", "Option"], decimals=2)
            .fmt_number(columns=["PCR", "OI_Chg_Pct", "Avg_IV", "V_OI"], decimals=2)
            .fmt_percent(columns=["Vol_Change", "OI_Change"], decimals=1)
            .data_color(columns=["Score"], palette=["red", "white", "green"], domain=[-100, 100])
            .data_color(columns=["Vol_Change"], palette=["red", "white", "lightgreen"], domain=[-50, 50])
            .data_color(columns=["OI_Change"], palette=["red", "white", "lightblue"], domain=[-50, 50])
            .data_color(columns=["PCR"], palette=["green", "white", "red"], domain=[0, 2])
            .tab_style(style=style.fill(color="yellow"), locations=loc.body(rows=lambda df: df['Is_New'] == 1))
            .cols_hide(columns=["Is_New"])
            .tab_options(table_font_size="12px", heading_background_color="#2E86AB", heading_text_color="white", stub_background_color="#F28E2C")
        )
        print("\n" + "="*120)
        if show_time: print(f"📊 {title} - {show_time}")
        else: print(f"📊 {title}")
        print("="*120)
        print("✨ Enhanced Great-Tables Display:")
        for i, row in df.iterrows():
            marker = "🆕 " if row['Is_New'] == 1 else "   "
            vol_chg = row['Vol_Change']
            oi_chg = row['OI_Change']
            vol_chg_str = f"{vol_chg:+.1f}%" if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1 else "N/A"
            oi_chg_str = f"{oi_chg:+.1f}%" if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1 else "N/A"
            color = Colors.MAGENTA if row['Is_New'] == 1 else Colors.END
            print_colored(f"{marker}{row['Stock']:<12} | {row['Signal']:<16} | {row['Score']:>7.2f} | {row['Curr_Vol']:>10} | {row['Curr_OI']:>10} | {vol_chg_str:>7} | {oi_chg_str:>7} | {row['PCR']:>5.2f} | {row['OI_Chg_Pct']:>7.2f} | {row['Avg_IV']:>6.2f} | {row['V_OI']:>5.2f} | {row['Option_Remark']:<12} | {row['Action']:<14}", color)
        print("="*120)
    except Exception as e:
        logger.error(f"Error creating great table: {e}")
        create_rich_enhanced_table(data, title, new_stocks, show_time)

def create_rich_enhanced_table(data, title, new_stocks=None, show_time=None):
    if not data:
        if RICH_AVAILABLE:
            console.print(f"\n[bold magenta]{title}[/bold magenta]")
            console.print("[yellow]No stocks found in this category.[/yellow]")
        else:
            print_colored(f"\n{title}", Colors.HEADER)
            print_colored("No stocks found in this category.", Colors.YELLOW)
        return
    if RICH_AVAILABLE:
        table = Table(box=box.ROUNDED, show_header=True, header_style="bold blue")
        table.add_column("Stock", style="bold white", width=12, justify="left")
        table.add_column("Signal", style="bold", width=16, justify="center")
        table.add_column("Score", style="bold", width=8, justify="right")
        table.add_column("Trend", style="cyan", width=7, justify="right")
        table.add_column("Momentum", style="yellow", width=9, justify="right")
        table.add_column("Volume", style="green", width=8, justify="right")
        table.add_column("OI", style="magenta", width=8, justify="right")
        table.add_column("Option", style="bright_blue", width=8, justify="right")  # New
        table.add_column("Curr Vol", style="bright_green", width=10, justify="right")
        table.add_column("Curr OI", style="bright_magenta", width=10, justify="right")
        table.add_column("Vol Δ%", style="bright_yellow", width=8, justify="right")
        table.add_column("OI Δ%", style="bright_cyan", width=8, justify="right")
        table.add_column("PCR", style="bright_red", width=6, justify="right")  # New
        table.add_column("Opt OI Δ%", style="bright_green", width=10, justify="right")  # New
        table.add_column("Avg IV", style="bright_yellow", width=8, justify="right")  # New
        table.add_column("V/OI", style="bright_cyan", width=6, justify="right")  # New
        table.add_column("Opt Remark", style="bold", width=12, justify="center")  # New
        table.add_column("Flow", style="dim", width=18, justify="left")
        table.add_column("Action", style="bold", width=14, justify="center")
        for item in data:
            symbol = item['symbol']
            is_new = new_stocks and symbol in new_stocks
            if item['score'] > 50: signal_style = "bold bright_green"
            elif item['score'] > 25: signal_style = "bold green"
            elif item['score'] > 0: signal_style = "green"
            elif item['score'] < -50: signal_style = "bold bright_red"
            elif item['score'] < -25: signal_style = "bold red"
            else: signal_style = "red"
            stock_style = "[bold bright_magenta]" + symbol + " ✨[/bold bright_magenta]" if is_new else symbol
            vol_change_raw = item.get('volume_change_pct', 0)
            if isinstance(vol_change_raw, str) or abs(vol_change_raw) < 0.1:
                vol_change_style = "[dim]N/A[/dim]"
            elif vol_change_raw > 20:
                vol_change_style = f"[bold bright_green]{vol_change_raw:+.1f}%[/bold bright_green]"
            elif vol_change_raw > 0:
                vol_change_style = f"[green]{vol_change_raw:+.1f}%[/green]"
            elif vol_change_raw < -20:
                vol_change_style = f"[bold bright_red]{vol_change_raw:+.1f}%[/bold bright_red]"
            else:
                vol_change_style = f"[red]{vol_change_raw:+.1f}%[/red]"
            oi_change_raw = item.get('oi_change_pct', 0)
            if isinstance(oi_change_raw, str) or abs(oi_change_raw) < 0.1:
                oi_change_style = "[dim]N/A[/dim]"
            elif oi_change_raw > 15:
                oi_change_style = f"[bold bright_cyan]{oi_change_raw:+.1f}%[/bold bright_cyan]"
            elif oi_change_raw > 0:
                oi_change_style = f"[cyan]{oi_change_raw:+.1f}%[/cyan]"
            elif oi_change_raw < -15:
                oi_change_style = f"[bold bright_red]{oi_change_raw:+.1f}%[/bold bright_red]"
            else:
                oi_change_style = f"[red]{oi_change_raw:+.1f}%[/red]"
            opt_oi_chg = item.get('oi_change_pct', 0)
            opt_oi_chg_style = f"[green]{opt_oi_chg:+.1f}%[/green]" if opt_oi_chg > 0 else f"[red]{opt_oi_chg:+.1f}%[/red]"
            pcr_raw = item.get('pcr', 0)
            pcr_style = f"[green]{pcr_raw:.2f}[/green]" if pcr_raw < 1 else f"[red]{pcr_raw:.2f}[/red]"
            remark_style = "green" if "Bullish" in item.get('option_remark', '') else "red" if "Bearish" in item.get('option_remark', '') else "yellow"
            table.add_row(
                stock_style,
                f"[{signal_style}]{item['signal']}[/{signal_style}]",
                f"[bold]{item['score']:.2f}[/bold]",
                f"{item['sub_scores'].get('Trend', 0):.2f}",
                f"{item['sub_scores'].get('Momentum', 0):.2f}",
                f"{item['sub_scores'].get('Volume', 0):.2f}",
                f"{item['sub_scores'].get('OI', 0):.2f}",
                f"{item['sub_scores'].get('Option', 0):.2f}",  # New
                f"[bright_green]{item.get('current_volume', 'N/A')}[/bright_green]",
                f"[bright_magenta]{item.get('current_oi', 'N/A')}[/bright_magenta]",
                vol_change_style,
                oi_change_style,
                pcr_style,  # New
                opt_oi_chg_style,  # New
                f"{item.get('avg_iv', 0):.2f}",  # New
                f"{item.get('v_oi_ratio', 0):.2f}",  # New
                f"[{remark_style}]{item.get('option_remark', 'Neutral')}[/{remark_style}]",  # New
                f"[dim]{item.get('flow', 'Unknown')}[/dim]",
                f"[bold]{item.get('action', 'Consider')}[/bold]"
            )
        if show_time:
            console.print(f"\n[bold magenta]{title} - {show_time}[/bold magenta]")
        else:
            console.print(f"\n[bold magenta]{title}[/bold magenta]")
        console.print(table)
    else:
        create_enhanced_ascii_table(data, title, new_stocks, show_time)

def create_enhanced_ascii_table(data, title, new_stocks=None, show_time=None):
    if not data:
        print_colored(f"\n{title}", Colors.HEADER)
        print_colored("No stocks found in this category.", Colors.YELLOW)
        return
    if show_time:
        print_colored(f"\n{title} - {show_time}", Colors.HEADER)
    else:
        print_colored(f"\n{title}", Colors.HEADER)
    print_colored("="*200, Colors.BLUE)
    header = (f"{'Stock':<12} | {'Signal':<16} | {'Score':>8} | {'Trend':>7} | {'Mom':>9} | {'Vol':>7} | {'OI':>7} | {'Option':>8} | "
              f"{'CurrVol':>10} | {'CurrOI':>10} | {'VolΔ%':>8} | {'OIΔ%':>8} | {'PCR':>5} | {'OptOIΔ%':>8} | {'AvgIV':>6} | {'V/OI':>5} | {'OptRemark':<12} | {'Flow':<18} | {'Action':<14}")
    print_colored(header, Colors.BOLD)
    print_colored("-"*200, Colors.BLUE)
    for item in data:
        symbol = item['symbol']
        is_new = new_stocks and symbol in new_stocks
        vol_chg = item.get('volume_change_pct', 0)
        oi_chg = item.get('oi_change_pct', 0)
        vol_chg_str = f"{vol_chg:+.1f}" if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1 else "N/A"
        oi_chg_str = f"{oi_chg:+.1f}" if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1 else "N/A"
        row = (f"{symbol:<12} | {item['signal']:<16} | {item['score']:>8.2f} | {item['sub_scores'].get('Trend', 0):>7.2f} | "
               f"{item['sub_scores'].get('Momentum', 0):>9.2f} | {item['sub_scores'].get('Volume', 0):>7.2f} | "
               f"{item['sub_scores'].get('OI', 0):>7.2f} | {item['sub_scores'].get('Option', 0):>8.2f} | {item.get('current_volume', 'N/A'):>10} | "
               f"{item.get('current_oi', 'N/A'):>10} | {vol_chg_str:>7}% | "
               f"{oi_chg_str:>7}% | {item.get('pcr', 'N/A'):>5.2f} | {item.get('oi_change_pct', 'N/A'):>7.2f} | "
               f"{item.get('avg_iv', 'N/A'):>6.2f} | {item.get('v_oi_ratio', 'N/A'):>5.2f} | {item.get('option_remark', 'Neutral'):<12} | "
               f"{item.get('flow', 'Unknown'):<18} | {item.get('action', 'Consider'):<14}")
        if is_new:
            print_colored(row + " ← ✨ NEW!", Colors.MAGENTA)
        else:
            print(row)
    print_colored("="*200, Colors.BLUE)

def create_compact_backtest_table(data, title, new_stocks=None, show_time=None):
    if not data:
        return
    if GREAT_TABLES_AVAILABLE:
        create_great_table_fixed(data[:Config.BACKTEST_TOP_DISPLAY], f"Compact {title}", new_stocks, show_time)
    elif RICH_AVAILABLE:
        table = Table(box=box.SIMPLE, show_header=True, header_style="bold blue")
        table.add_column("#", width=3, justify="right")
        table.add_column("Stock", style="bold white", width=12)
        table.add_column("Signal", style="bold", width=14)
        table.add_column("Score", style="bold", width=8, justify="right")
        table.add_column("CurrVol", style="bright_green", width=8, justify="right")
        table.add_column("CurrOI", style="bright_magenta", width=8, justify="right")
        table.add_column("VolΔ%", style="bright_yellow", width=7, justify="right")
        table.add_column("OIΔ%", style="bright_cyan", width=7, justify="right")
        table.add_column("PCR", style="bright_red", width=6, justify="right")  # New
        table.add_column("Opt Remark", style="bold", width=12)  # New
        table.add_column("Action", style="bold", width=12)
        for i, item in enumerate(data, 1):
            symbol = item['symbol']
            is_new = new_stocks and symbol in new_stocks
            stock_display = f"[bright_magenta]{symbol} ✨[/bright_magenta]" if is_new else symbol
            vol_chg = item.get('volume_change_pct', 0)
            if isinstance(vol_chg, str) or abs(vol_chg) < 0.1:
                vol_display = "[dim]N/A[/dim]"
            elif vol_chg > 0:
                vol_display = f"[bright_green]{vol_chg:+.1f}%[/bright_green]"
            else:
                vol_display = f"[red]{vol_chg:+.1f}%[/red]"
            oi_chg = item.get('oi_change_pct', 0)
            if isinstance(oi_chg, str) or abs(oi_chg) < 0.1:
                oi_display = "[dim]N/A[/dim]"
            elif oi_chg > 0:
                oi_display = f"[bright_cyan]{oi_chg:+.1f}%[/bright_cyan]"
            else:
                oi_display = f"[red]{oi_chg:+.1f}%[/red]"
            table.add_row(
                str(i),
                stock_display,
                item['signal'],
                f"{item['score']:.1f}",
                str(item.get('current_volume', 'N/A')),
                str(item.get('current_oi', 'N/A')),
                vol_display,
                oi_display,
                f"{item.get('pcr', 'N/A'):.2f}",  # New
                item.get('option_remark', 'Neutral'),  # New
                item.get('action', 'Consider')
            )
        if show_time:
            console.print(f"\n[bold blue]{title} - {show_time}[/bold blue]")
        else:
            console.print(f"\n[bold blue]{title}[/bold blue]")
        console.print(table)
    else:
        create_enhanced_ascii_table(data[:Config.BACKTEST_TOP_DISPLAY], title, new_stocks, show_time)

# ========== TECHNICAL INDICATORS ==========
def ema(series, length):
    return series.ewm(span=length, adjust=False).mean()

def vwap(df, period=None):
    price = (df["High"] + df["Low"] + df["Close"]) / 3.0
    pv = price * df["Volume"]
    if period:
        pv_sum = pv.rolling(period).sum()
        vol_sum = df["Volume"].rolling(period).sum()
    else:
        pv_sum = pv.cumsum()
        vol_sum = df["Volume"].cumsum()
    return pv_sum / vol_sum.replace(0, np.nan)

def atr(df, period=14):
    high_low = df["High"] - df["Low"]
    high_close = (df["High"] - df["Close"].shift(1)).abs()
    low_close = (df["Low"] - df["Close"].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    return tr.ewm(alpha=1/period, adjust=False).mean()

def williams_r(df, period=14):
    highest = df["High"].rolling(period).max()
    lowest = df["Low"].rolling(period).min()
    return -100 * (highest - df["Close"]) / (highest - lowest).replace(0, np.nan)

def momentum(df, period=10):
    return df["Close"] / df["Close"].shift(period) - 1.0

def volume_surge(df, lookback=20):
    vol_ma = df["Volume"].rolling(lookback).mean()
    vol_std = df["Volume"].rolling(lookback).std()
    z_score = (df["Volume"] - vol_ma) / vol_std.replace(0, np.nan)
    return z_score.fillna(0)

def calculate_rsi(df, period=14):
    if len(df) < period + 1: 
        return pd.Series(dtype='float64', index=df.index)
    delta = df['Close'].diff()
    gain = (delta.where(delta > 0, 0)).ewm(com=period - 1, adjust=False).mean()
    loss = (-delta.where(delta < 0, 0)).ewm(com=period - 1, adjust=False).mean()
    rs = gain / loss.replace(0, np.nan)
    rs.fillna(100, inplace=True)
    return 100 - (100 / (1 + rs))

def calculate_macd(df, fast=12, slow=26, signal=9):
    if len(df) < slow + signal: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    ema_fast = df['Close'].ewm(span=fast, adjust=False).mean()
    ema_slow = df['Close'].ewm(span=slow, adjust=False).mean()
    macd = ema_fast - ema_slow
    signal_line = macd.ewm(span=signal, adjust=False).mean()
    return macd, signal_line

def calculate_stochastic(df, period=14, smooth_d=3):
    if len(df) < period + smooth_d: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    low_min = df['Low'].rolling(window=period).min()
    high_max = df['High'].rolling(window=period).max()
    k = 100 * ((df['Close'] - low_min) / (high_max - low_min).replace(0, np.nan))
    k.fillna(50, inplace=True)
    d = k.rolling(window=smooth_d).mean()
    return k, d
    
def calculate_adx(df, period=14):
    if len(df) < period * 2: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    df_adx = df.copy()
    df_adx['H-L'] = df_adx['High'] - df_adx['Low']
    df_adx['H-C'] = abs(df_adx['High'] - df_adx['Close'].shift(1))
    df_adx['L-C'] = abs(df_adx['Low'] - df_adx['Close'].shift(1))
    df_adx['TR'] = df_adx[['H-L', 'H-C', 'L-C']].max(axis=1)
    df_adx['+DM'] = np.where((df_adx['High'] - df_adx['High'].shift(1)) > (df_adx['Low'].shift(1) - df_adx['Low']), df_adx['High'] - df_adx['High'].shift(1), 0)
    df_adx['-DM'] = np.where((df_adx['Low'].shift(1) - df_adx['Low']) > (df_adx['High'] - df_adx['High'].shift(1)), df_adx['Low'].shift(1) - df_adx['Low'], 0)
    atr_val = df_adx['TR'].ewm(com=period - 1, adjust=False).mean().replace(0, np.nan)
    pdi = (df_adx['+DM'].ewm(com=period - 1, adjust=False).mean() / atr_val) * 100
    ndi = (df_adx['-DM'].ewm(com=period - 1, adjust=False).mean() / atr_val) * 100
    adx = (abs(pdi - ndi) / (pdi + ndi).replace(0, np.nan)).ewm(com=period - 1, adjust=False).mean() * 100
    return adx.fillna(20), pdi.fillna(20), ndi.fillna(20)

def calculate_bollinger_bands(df, period=20, std_dev=2):
    if len(df) < period:
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    middle = df['Close'].rolling(window=period).mean()
    std = df['Close'].rolling(window=period).std()
    upper = middle + (std * std_dev)
    lower = middle - (std * std_dev)
    return middle, upper, lower

def calculate_roc(df, period=12):
    if len(df) < period + 1:
        return pd.Series(dtype='float64', index=df.index)
    return (df['Close'] / df['Close'].shift(period) - 1) * 100

def calculate_cci(df, period=20):
    if len(df) < period:
        return pd.Series(dtype='float64', index=df.index)
    tp = (df['High'] + df['Low'] + df['Close']) / 3
    tp_ma = tp.rolling(period).mean()
    tp_md = tp.rolling(period).apply(lambda x: np.abs(x - x.mean()).mean())
    return (tp - tp_ma) / (0.015 * tp_md).replace(0, np.nan).fillna(0)

def calculate_obv(df):
    obv = (np.sign(df['Close'].diff()) * df['Volume']).fillna(0).cumsum()
    return obv

def calculate_cmf(df, period=20):
    if len(df) < period:
        return pd.Series(dtype='float64', index=df.index)
    mfv = ((df['Close'] - df['Low'] - (df['High'] - df['Close'])) / (df['High'] - df['Low']).replace(0, np.nan)) * df['Volume']
    mfv.fillna(0, inplace=True)
    return mfv.rolling(period).sum() / df['Volume'].rolling(period).sum().replace(0, np.nan).fillna(0)

def calculate_relative_volume(df, lookback=20):
    return df['Volume'] / df['Volume'].rolling(lookback).mean().replace(0, np.nan).fillna(1)

# ========== INDICATOR SCORING ==========
def get_indicator_score(ind_name, value, df=None, period=None):
    if ind_name == "MA_Slope":
        return value * 10  # Positive slope bullish
    elif ind_name == "ADX":
        return min(value / 25 * 5, 5) if value > 25 else 0
    elif ind_name == "VWAP":
        return 5 if df['Close'].iloc[-1] > value.iloc[-1] else -5
    elif ind_name == "EMA":
        return 5 if df['Close'].iloc[-1] > value.iloc[-1] else -5
    elif ind_name == "MACD_Trend":
        return value * 5
    elif ind_name == "RSI":
        if value > 70: return -10
        elif value < 30: return 10
        else: return (50 - value) / 2
    elif ind_name == "Stochastic":
        if value > 80: return -5
        elif value < 20: return 5
        else: return 0
    elif ind_name == "CCI":
        if value > 100: return -5
        elif value < -100: return 5
        else: return 0
    elif ind_name == "ROC":
        return min(max(value, -20), 20)
    elif ind_name == "WilliamsR":
        if value < -80: return 5
        elif value > -20: return -5
        else: return 0
    elif ind_name == "VolumeSurge":
        return min(value * 2, 10)
    elif ind_name == "OBV":
        return 5 if value > 0 else -5
    elif ind_name == "CMF":
        return value * 10
    elif ind_name == "RelVol":
        return min((value - 1) * 5, 10)
    elif ind_name == "VolatilityExpansion":
        return value * 5
    elif ind_name == "Bollinger":
        return value * 5
    elif ind_name == "OptionBuyerMomentum":
        return value * 10
    elif ind_name == "OIChange":
        return value * 5
    elif ind_name == "VolumeOISync":
        return value * 5
    return 0

# ========== CALCULATE INDICATORS ==========
def calculate_multi_timeframe_indicators(data):
    indicators = {}
    for tf, df in data.items():
        if df.empty: continue
        indicators[tf] = {}
        # Trend
        indicators[tf]['MA_Slope'] = (df['Close'].ewm(span=50).mean() - df['Close'].ewm(span=200).mean()) / df['Close'].ewm(span=200).mean()
        adx, pdi, ndi = calculate_adx(df)
        indicators[tf]['ADX'] = adx.iloc[-1]
        indicators[tf]['VWAP'] = vwap(df).iloc[-1]
        indicators[tf]['EMA'] = ema(df['Close'], 20).iloc[-1]
        macd, signal = calculate_macd(df)
        indicators[tf]['MACD_Trend'] = 1 if macd.iloc[-1] > signal.iloc[-1] else -1
        # Momentum
        indicators[tf]['RSI'] = calculate_rsi(df).iloc[-1]
        k, d = calculate_stochastic(df)
        indicators[tf]['Stochastic'] = k.iloc[-1]
        indicators[tf]['CCI'] = calculate_cci(df).iloc[-1]
        indicators[tf]['ROC'] = calculate_roc(df).iloc[-1]
        indicators[tf]['WilliamsR'] = williams_r(df).iloc[-1]
        # Volume
        indicators[tf]['VolumeSurge'] = volume_surge(df).iloc[-1]
        indicators[tf]['OBV'] = calculate_obv(df).iloc[-1]
        indicators[tf]['CMF'] = calculate_cmf(df).iloc[-1]
        indicators[tf]['RelVol'] = calculate_relative_volume(df).iloc[-1]
        # Volatility
        indicators[tf]['VolatilityExpansion'] = atr(df).iloc[-1] / df['Close'].iloc[-1]
        middle, upper, lower = calculate_bollinger_bands(df)
        indicators[tf]['Bollinger'] = (df['Close'].iloc[-1] - lower.iloc[-1]) / (upper.iloc[-1] - lower.iloc[-1]) - 0.5
        # OI
        if 'OpenInterest' in df.columns:
            indicators[tf]['OptionBuyerMomentum'] = momentum(df, 5).iloc[-1] if 'OpenInterest' in df else 0
            indicators[tf]['OIChange'] = (df['OpenInterest'].pct_change() * 100).iloc[-1]
            indicators[tf]['VolumeOISync'] = 1 if df['Volume'].iloc[-1] > df['Volume'].shift(1).iloc[-1] and df['OpenInterest'].iloc[-1] > df['OpenInterest'].shift(1).iloc[-1] else -1
    return indicators

# ========== CALCULATE SCORE (UPDATED WITH OPTION CHAIN) ==========
def calculate_score(multi_tf_indicators, market_regime, option_metrics={}):
    sub_scores = {group: 0 for group in Config.GROUP_WEIGHTS}
    score = 0
    for tf, inds in multi_tf_indicators.items():
        tf_weight = TIMEFRAME_WEIGHTS.get(tf, 1.0)
        for group, group_inds in {
            "Trend": ["MA_Slope", "ADX", "VWAP", "EMA", "MACD_Trend"],
            "Momentum": ["RSI", "Stochastic", "CCI", "ROC", "WilliamsR"],
            "Volume": ["VolumeSurge", "OBV", "CMF", "RelVol"],
            "Volatility": ["VolatilityExpansion", "Bollinger"],
            "OI": ["OptionBuyerMomentum", "OIChange", "VolumeOISync"]
        }.items():
            group_score = 0
            for ind in group_inds:
                value = inds.get(ind, 0)
                ind_score = get_indicator_score(ind, value)
                group_score += ind_score * Config.INDICATOR_WEIGHTS.get(ind, 1.0)
            sub_scores[group] += group_score * tf_weight / len(multi_tf_indicators)

    # Add Option group score based on remark
    option_sentiment = 0
    option_remark = option_metrics.get('Remark', 'Neutral')
    if option_remark == 'Strong Bullish':
        option_sentiment = 20
    elif option_remark == 'Mild Bullish':
        option_sentiment = 10
    elif option_remark == 'Strong Bearish':
        option_sentiment = -20
    elif option_remark == 'Mild Bearish':
        option_sentiment = -10
    sub_scores['Option'] = option_sentiment * Config.INDICATOR_WEIGHTS.get("OptionSentiment", 1.0)

    # Aggregate
    for group, group_score in sub_scores.items():
        score += group_score * Config.GROUP_WEIGHTS.get(group, 1.0)

    # Apply regime multiplier
    is_bullish = score > 0
    regime_mult = Config.REGIME_MULTIPLIERS.get(
        'bullish_in_bull_market' if is_bullish and market_regime == 'bullish' else
        'bearish_in_bear_market' if not is_bullish and market_regime == 'bearish' else
        'bullish_in_bear_market' if is_bullish and market_regime == 'bearish' else
        'bearish_in_bull_market'
    )
    score *= regime_mult

    return score, sub_scores

# ========== PROCESS STOCK ==========
def process_stock(symbol, data, option_metrics={}, market_regime='neutral'):
    if not data or not data.get(5): return None
    multi_tf_indicators = calculate_multi_timeframe_indicators(data)
    score, sub_scores = calculate_score(multi_tf_indicators, market_regime, option_metrics)
    if abs(score) < Config.SCORE_THRESHOLD_MIN: return None

    current_volume = data[5]['Volume'].iloc[-1] if 'Volume' in data[5] else 'N/A'
    current_oi = data[5]['OpenInterest'].iloc[-1] if 'OpenInterest' in data[5] else 'N/A'
    volume_change_pct = (data[5]['Volume'].pct_change() * 100).iloc[-1] if len(data[5]) > 1 else 0
    oi_change_pct = (data[5]['OpenInterest'].pct_change() * 100).iloc[-1] if len(data[5]) > 1 and 'OpenInterest' in data[5] else 0

    flow = 'Bullish Flow' if score > 0 else 'Bearish Flow'
    action = 'Buy Call' if score > 0 else 'Buy Put'

    thresholds = sorted(Config.SIGNAL_THRESHOLDS.items(), key=lambda x: x[1], reverse=score > 0)
    signal = 'Neutral'
    for name, thresh in thresholds:
        if (score > 0 and score >= thresh) or (score < 0 and score <= thresh):
            signal = name
            break

    result = {
        'symbol': symbol,
        'score': score,
        'sub_scores': sub_scores,
        'signal': signal,
        'current_volume': current_volume,
        'current_oi': current_oi,
        'volume_change_pct': volume_change_pct,
        'oi_change_pct': oi_change_pct,
        'flow': flow,
        'action': action,
        'pcr': option_metrics.get('PCR', 'N/A'),
        'oi_change_pct': option_metrics.get('OI Chg %', 'N/A'),
        'avg_iv': option_metrics.get('Avg IV %', 'N/A'),
        'v_oi_ratio': option_metrics.get('V/OI Ratio', 'N/A'),
        'atm_signal': option_metrics.get('ATM Signal', 'N/A'),
        'option_remark': option_metrics.get('Remark', 'Neutral')
    }
    return result

# ========== FETCH PARALLEL ==========
def fetch_parallel(stocks, asof_ts, is_live):
    data_dict = {}
    with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_hist, symbol, asof_ts, is_live): symbol for symbol in stocks}
        for future in as_completed(futures):
            symbol = futures[future]
            try:
                data_dict[symbol] = future.result()
            except Exception as e:
                logger.error(f"Fetch error for {symbol}: {e}")
    return data_dict

def fetch_hist(symbol, asof_ts, is_live):
    sess = random.choice(td_pool)
    data = {}
    for tf in [5, 15, 30, 60, 1440]:
        bar_size = BAR_SIZE_MAP[tf]
        duration = DURATION_MAP[tf]
        end_time = asof_ts if not is_live else datetime.now(IST)
        start_time = end_time - timedelta(days=int(duration.split(' ')[0]))
        df = sess.get_hist(symbol, bar_size=bar_size, duration=duration, end=end_time, start=start_time)
        if 'oi' in df.columns:
            df.rename(columns={'oi': 'OpenInterest'}, inplace=True)
        data[tf] = df
    return data

# ========== RUN SCAN AT TIME (UPDATED WITH OPTION FETCH) ==========
def run_scan_at_time_5min_fixed(scan_time, stocks, market_regime, is_live=False):
    data_dict = fetch_parallel(stocks, scan_time, is_live)

    # Parallel fetch option chain
    option_dict = {}
    with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_symbol_metrics, symbol.split('-')[0]): symbol for symbol in stocks}
        for future in as_completed(futures):
            symbol = futures[future]
            try:
                option_dict[symbol] = future.result()
            except Exception as e:
                logger.error(f"Option fetch error for {symbol}: {e}")

    signals = []
    current_symbols = set()
    with ThreadPoolExecutor(max_workers=Config.MAX_WORKERS) as executor:
        futures = {executor.submit(process_stock, symbol, data_dict.get(symbol, {}), option_dict.get(symbol, {}), market_regime): symbol for symbol in stocks}
        for future in as_completed(futures):
            result = future.result()
            if result:
                signals.append(result)
                current_symbols.add(result['symbol'])

    return signals, current_symbols

# ========== GET MARKET REGIME ==========
def get_market_regime(benchmark):
    # Placeholder for fetching benchmark data
    # In practice, fetch NIFTY 50 data and determine if above/below MA
    return 'bullish'  # or 'bearish'

# ========== SLEEP UNTIL ==========
def sleep_until(target_time):
    now = datetime.now(IST)
    sleep_sec = (target_time - now).total_seconds()
    if sleep_sec > 0:
        time.sleep(sleep_sec)

# ========== GET EXACT CANDLE CLOSE TIME ==========
def get_exact_candle_close_time(now):
    minute = now.minute
    to_next = (4 - (minute % 5)) % 5
    if to_next == 0:
        to_next = 5
    next_close = now + timedelta(minutes=to_next)
    next_close = next_close.replace(second=0, microsecond=0)
    return next_close + timedelta(seconds=Config.SETTLE_DELAY_SECONDS)

def today_ist_dt(time_str):
    today = datetime.now(IST).date()
    t = datetime.strptime(time_str, '%H:%M').time()
    return IST.localize(datetime.combine(today, t))

# ========== RUN FULL DAY BACKTEST ==========
def run_full_day_backtest_5min_fixed(backtest_date, stocks):
    try:
        backtest_dt = datetime.strptime(backtest_date, "%Y-%m-%d")
    except ValueError:
        logger.error("Invalid date format. Use YYYY-MM-DD.")
        return

    logger.info(f"Starting FIXED 5-min backtest for {backtest_date}")

    market_start = datetime.strptime(Config.MARKET_START, "%H:%M").time()
    first_run = datetime.strptime(Config.FIRST_RUN_AT, "%H:%M").time()
    market_end = datetime.strptime(Config.MARKET_END, "%H:%M").time()

    start_dt = IST.localize(datetime.combine(backtest_dt, market_start)) + timedelta(minutes=Config.BACKTEST_START_DELAY)
    end_dt = IST.localize(datetime.combine(backtest_dt, market_end))

    scan_times = []
    current = start_dt
    while current < end_dt:
        scan_times.append(current)
        current += timedelta(minutes=Config.BACKTEST_INTERVAL_MINUTES)

    total_scans = len(scan_times)
    all_results = []
    market_regime = get_market_regime(Config.BENCHMARK_INDEX)
    global previous_scan_results
    previous_scan_results = {}

    with tqdm(total=total_scans, desc="Backtesting Progress", unit="scan") as pbar:
        for i, scan_time in enumerate(scan_times):
            try:
                signals, current_symbols = run_scan_at_time_5min_fixed(scan_time, stocks, market_regime, is_live=False)
                new_stocks = current_symbols - set(previous_scan_results.keys()) if previous_scan_results else current_symbols
                previous_scan_results = {s: True for s in current_symbols}

                timestamp = scan_time.isoformat()
                scan_result = {
                    'timestamp': timestamp,
                    'total_signals': len(signals),
                    'bullish_signals': len([s for s in signals if s['score'] > 0]),
                    'bearish_signals': len([s for s in signals if s['score'] < 0]),
                    'new_stocks': list(new_stocks),
                    'signals': signals
                }
                all_results.append(scan_result)
                
                if signals:
                    signals.sort(key=lambda x: abs(x['score']), reverse=True)
                    top_bullish = [r for r in signals if r['score'] > 0][:Config.BACKTEST_TOP_DISPLAY]
                    top_bearish = [r for r in signals if r['score'] < 0][:Config.BACKTEST_TOP_DISPLAY]
                    
                    scan_time_str = scan_time.strftime('%H:%M')
                    
                    # Count stocks with meaningful volume/OI changes
                    vol_with_changes = sum(1 for s in signals if isinstance(s.get('volume_change_pct', 0), (int, float)) and abs(s.get('volume_change_pct', 0)) > 0.1)
                    oi_with_changes = sum(1 for s in signals if isinstance(s.get('oi_change_pct', 0), (int, float)) and abs(s.get('oi_change_pct', 0)) > 0.1)
                    
                    if RICH_AVAILABLE:
                        console.print(f"\n[bold blue]🔥 SCAN #{i+1}/{total_scans} - {scan_time_str} IST[/bold blue]")
                        console.print(f"[cyan]Signals: {len(signals)} | Bullish: {len([s for s in signals if s['score'] > 0])} | Bearish: {len([s for s in signals if s['score'] < 0])} | New: {len(new_stocks)}[/cyan]")
                        console.print(f"[yellow]Volume Changes: {vol_with_changes} stocks | OI Changes: {oi_with_changes} stocks[/yellow]")
                    else:
                        print_colored(f"\n🔥 SCAN #{i+1}/{total_scans} - {scan_time_str} IST", Colors.BOLD)
                        print_colored(f"Signals: {len(signals)} | Bullish: {len([s for s in signals if s['score'] > 0])} | Bearish: {len([s for s in signals if s['score'] < 0])} | New: {len(new_stocks)}", Colors.CYAN)
                        print_colored(f"Volume Changes: {vol_with_changes} stocks | OI Changes: {oi_with_changes} stocks", Colors.YELLOW)
                    
                    if top_bullish:
                        if GREAT_TABLES_AVAILABLE:
                            create_great_table_fixed(top_bullish, "🟢 TOP BULLISH", new_stocks, scan_time_str)
                        else:
                            create_compact_backtest_table(top_bullish, "🟢 TOP BULLISH", new_stocks, scan_time_str)
                    
                    if top_bearish:
                        if GREAT_TABLES_AVAILABLE:
                            create_great_table_fixed(top_bearish, "🔴 TOP BEARISH", new_stocks, scan_time_str)
                        else:
                            create_compact_backtest_table(top_bearish, "🔴 TOP BEARISH", new_stocks, scan_time_str)
                    
                    if new_stocks and len(new_stocks) > 0:
                        new_stocks_display = list(new_stocks)[:10]
                        more_text = f"... +{len(new_stocks)-10}" if len(new_stocks) > 10 else ""
                        print_colored(f"\n✨ NEW STOCKS: {', '.join(new_stocks_display)}{more_text}", Colors.MAGENTA)
                        
                else:
                    print_colored(f"\n[{scan_time.strftime('%H:%M')}] Scan #{i+1}/{total_scans} - No signals", Colors.YELLOW)
                
                pbar.update(1)
                time.sleep(0.1)
                    
            except Exception as e:
                logger.error(f"Error in backtest scan at {scan_time}: {e}")
                pbar.update(1)
                continue
    
    # Show enhanced summary
    print_colored(f"\n📊 FIXED 5-MIN BACKTEST SUMMARY FOR {backtest_date}", Colors.HEADER)
    print_colored("="*120, Colors.BLUE)
    
    total_scans_completed = len([r for r in all_results if r['total_signals'] >= 0])
    total_signals = sum(r['total_signals'] for r in all_results)
    total_bullish = sum(r['bullish_signals'] for r in all_results)
    total_bearish = sum(r['bearish_signals'] for r in all_results)
    unique_stocks = len(set(stock for r in all_results for stock in [s['symbol'] for s in r['signals']]))
    
    print(f"📈 Scans: {total_scans_completed}/{total_scans}")
    print(f"📊 Total Signals: {total_signals}")
    print(f"🟢 Bullish: {total_bullish}")
    print(f"🔴 Bearish: {total_bearish}")
    print(f"📋 Unique Stocks: {unique_stocks}")
    
    if total_signals > 0:
        print(f"📊 Avg Signals/Scan: {total_signals/total_scans_completed:.1f}")
        print(f"⚖️  Bull/Bear Ratio: {total_bullish/max(total_bearish, 1):.2f}")
    
    # FIXED: Enhanced Volume/OI change statistics
    print_colored("\n📊 5-MINUTE VOLUME/OI CHANGE STATISTICS:", Colors.CYAN)
    vol_changes = []
    oi_changes = []
    
    for result in all_results[1:]:  # Skip first scan
        for signal in result['signals']:
            vol_chg = signal.get('volume_change_pct', 0)
            oi_chg = signal.get('oi_change_pct', 0)
            if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1:
                vol_changes.append(vol_chg)
            if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1:
                oi_changes.append(oi_chg)
    
    if vol_changes:
        avg_vol_chg = sum(vol_changes) / len(vol_changes)
        max_vol_chg = max(vol_changes)
        min_vol_chg = min(vol_changes)
        print(f"  📈 Volume Changes: {len(vol_changes)} stocks with meaningful changes")
        print(f"     Average: {avg_vol_chg:.1f}% | Max: {max_vol_chg:.1f}% | Min: {min_vol_chg:.1f}%")
    else:
        print(f"  📈 Volume Changes: No meaningful changes detected (threshold: >0.1%)")
    
    if oi_changes:
        avg_oi_chg = sum(oi_changes) / len(oi_changes)
        max_oi_chg = max(oi_changes)
        min_oi_chg = min(oi_changes)
        print(f"  📊 OI Changes: {len(oi_changes)} stocks with meaningful changes")
        print(f"     Average: {avg_oi_chg:.1f}% | Max: {max_oi_chg:.1f}% | Min: {min_oi_chg:.1f}%")
    else:
        print(f"  📊 OI Changes: No meaningful changes detected (threshold: >0.1%)")
    
    # Most active times
    active_scans = sorted(all_results, key=lambda x: x['total_signals'], reverse=True)[:5]
    print_colored("\n🔥 MOST ACTIVE TIMES:", Colors.CYAN)
    for i, scan in enumerate(active_scans):
        if scan['total_signals'] > 0:
            time_str = datetime.fromisoformat(scan['timestamp']).strftime('%H:%M')
            vol_active = sum(1 for s in scan['signals'] if isinstance(s.get('volume_change_pct', 0), (int, float)) and abs(s.get('volume_change_pct', 0)) > 0.1)
            print(f"  {i+1}. {time_str} - {scan['total_signals']} signals ({scan['bullish_signals']}B/{scan['bearish_signals']}S) | {vol_active} vol changes")
    
    # Save results
    output_filename = f"{backtest_date}_5min_fixed_backtest_results.json"
    try:
        with open(output_filename, 'w') as f:
            json.dump(all_results, f, indent=2)
        print_colored(f"\n💾 Results saved: {output_filename}", Colors.GREEN)
    except Exception as e:
        logger.error(f"Could not save results: {e}")
    
    print_colored("="*120, Colors.BLUE)
    print_colored("🎯 Fixed 5-minute backtesting completed!", Colors.GREEN)

# ========== ADDITIONAL UTILITY FUNCTIONS ==========
def export_backtest_to_excel(results, filename):
    """Export backtest results to Excel with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Scans", len(results)],
            ["Total Signals", sum(r['total_signals'] for r in results)],
            ["Total Bullish", sum(r['bullish_signals'] for r in results)],
            ["Total Bearish", sum(r['bearish_signals'] for r in results)],
            ["Unique Stocks", len(set(stock for r in results for stock in r.get('new_stocks', [])))],
        ]
        
        for row_idx, (metric, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Signals sheet
        ws_signals = wb.create_sheet("All Signals")
        headers = ["Timestamp", "Stock", "Signal", "Score", "Action", "Volume Change", "OI Change"]
        
        for col_idx, header in enumerate(headers, 1):
            ws_signals.cell(row=1, column=col_idx, value=header)
        
        row_idx = 2
        for result in results:
            for signal in result.get('signals', []):
                ws_signals.cell(row=row_idx, column=1, value=result['timestamp'])
                ws_signals.cell(row=row_idx, column=2, value=signal['symbol'])
                ws_signals.cell(row=row_idx, column=3, value=signal['signal'])
                ws_signals.cell(row=row_idx, column=4, value=signal['score'])
                ws_signals.cell(row=row_idx, column=5, value=signal.get('action', 'N/A'))
                ws_signals.cell(row=row_idx, column=6, value=signal.get('volume_change_pct', 'N/A'))
                ws_signals.cell(row=row_idx, column=7, value=signal.get('oi_change_pct', 'N/A'))
                row_idx += 1
        
        wb.save(filename)
        print_colored(f"📊 Excel export saved: {filename}", Colors.GREEN)
        
    except ImportError:
        print_colored("⚠️  openpyxl not installed. Install with: pip install openpyxl", Colors.YELLOW)
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")

def generate_performance_report(results, backtest_date):
    """Generate a comprehensive performance report."""
    print_colored(f"\n📈 COMPREHENSIVE PERFORMANCE REPORT - {backtest_date}", Colors.HEADER)
    print_colored("="*100, Colors.BLUE)
    
    if not results:
        print_colored("No data available for report generation.", Colors.YELLOW)
        return
    
    # Time-based analysis
    hourly_signals = defaultdict(int)
    for result in results:
        timestamp = datetime.fromisoformat(result['timestamp'])
        hour = timestamp.hour
        hourly_signals[hour] += result['total_signals']
    
    print_colored("\n⏰ HOURLY ACTIVITY BREAKDOWN:", Colors.CYAN)
    for hour in sorted(hourly_signals.keys()):
        bar_length = min(50, int(hourly_signals[hour] * 50 / max(hourly_signals.values(), default=1)))
        bar = "█" * bar_length
        print(f"  {hour:02d}:00 | {hourly_signals[hour]:4d} signals | {bar}")
    
    # Signal strength distribution
    strong_signals = sum(len([s for s in r['signals'] if 'Strong' in s.get('signal', '')]) for r in results)
    total_signals = sum(r['total_signals'] for r in results)
    
    print_colored(f"\n💪 SIGNAL STRENGTH ANALYSIS:", Colors.CYAN)
    print(f"  Strong Signals: {strong_signals}/{total_signals} ({strong_signals/max(total_signals,1)*100:.1f}%)")
    
    # Volume/OI change analysis
    significant_vol_changes = 0
    significant_oi_changes = 0
    
    for result in results:
        for signal in result.get('signals', []):
            vol_chg = signal.get('volume_change_pct', 0)
            oi_chg = signal.get('oi_change_pct', 0)
            
            if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 20:
                significant_vol_changes += 1
            if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 15:
                significant_oi_changes += 1
    
    print_colored(f"\n📊 SIGNIFICANT CHANGES:", Colors.CYAN)
    print(f"  Volume Changes >20%: {significant_vol_changes}")
    print(f"  OI Changes >15%: {significant_oi_changes}")
    
    # Export option
    export_filename = f"{backtest_date}_performance_report.xlsx"
    try:
        export_backtest_to_excel(results, export_filename)
    except Exception as e:
        logger.error(f"Could not create Excel export: {e}")
    
    print_colored("="*100, Colors.BLUE)

# ========== MAIN EXECUTION WITH ENHANCED ERROR HANDLING ==========
if __name__ == "__main__":
    try:
        # Display startup banner
        print_colored("\n🎯 MERGED ENHANCED OPTION BUYER SCANNER v4.0", Colors.HEADER)
        print_colored("✅ Integrated TrueData TA + Option Chain Analysis", Colors.GREEN)
        
        if GREAT_TABLES_AVAILABLE:
            print_colored("✨ Great-Tables: Available for beautiful visualizations", Colors.GREEN)
        elif RICH_AVAILABLE:
            print_colored("✨ Rich: Available for enhanced tables", Colors.GREEN)
        else:
            print_colored("ℹ️  ASCII: Using fallback table formatting", Colors.YELLOW)
        
        print_colored("\n🔧 FEATURES:", Colors.CYAN)
        print("  🎯 Option Chain PCR, OI%, IV integrated into scoring")
        print("  📈 Multi-timeframe TA with weighted groups")
        print("  🎨 Enhanced tables with option metrics")
        print("  📊 Backtesting with statistics and export")
        print("  🔄 Parallel data fetching for speed")
        
        print_colored("\n📋 USAGE EXAMPLES:", Colors.YELLOW)
        print("  🔴 Live Trading:     python merger_grok.py")
        print("  🔍 Single Snapshot:  python merger_grok.py --asof 2025-09-30T14:25")
        print("  📈 Full Day Backtest: python merger_grok.py --backtest 2025-09-30")
        
        print_colored("="*90, Colors.HEADER)
        
        # Install recommendations
        if not GREAT_TABLES_AVAILABLE:
            print_colored("\n💡 RECOMMENDATION: Install great-tables for best visualization:", Colors.CYAN)
            print("   pip install great-tables")
        
        if not RICH_AVAILABLE:
            print_colored("\n💡 RECOMMENDATION: Install rich for enhanced output:", Colors.CYAN)
            print("   pip install rich")
        
        print_colored("\n🎯 Starting Merged Scanner...", Colors.GREEN)
        
        # Run main function
        main_final_fixed()
        
    except KeyboardInterrupt:
        print_colored("\n\n⚠️  Scanner interrupted by user. Shutting down gracefully...", Colors.YELLOW)
        
        # Show final statistics if available
        if 'scan_count' in globals() and scan_count > 0:
            print_colored(f"📊 Total scans completed: {scan_count}", Colors.CYAN)
        
    except ImportError as e:
        print_colored(f"\n❌ Import Error: {e}", Colors.RED)
        print_colored("💡 Please install required packages:", Colors.YELLOW)
        print("   pip install pandas numpy tqdm rich colorama great-tables requests truedata-ws openpyxl")
        
    except Exception as e:
        logger.exception(f"❌ Fatal error occurred: {e}")
        print_colored(f"\n💥 Unexpected error: {e}", Colors.RED)
        print_colored("📋 Please check your configuration and try again.", Colors.YELLOW)
        raise
        
    finally:
        # Cleanup and shutdown
        print_colored("\n🔌 Cleaning up resources...", Colors.CYAN)
        
        try:
            # Disconnect TrueData sessions (updated)
            for i, sess in enumerate(td_pool):
                try:
                    sess.disconnect()
                except Exception as cleanup_error:
                    logger.error(f"Error disconnecting session {i}: {cleanup_error}")
            
            print_colored("✅ All TrueData sessions disconnected.", Colors.GREEN)
            
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Final message
        print_colored("\n🎯 Merged Enhanced Option Buyer Scanner v4.0 shutdown complete!", Colors.HEADER)
        print_colored("📊 Thank you for using the Merged Scanner with Option Chain Integration!", Colors.GREEN)
        
        # Performance summary if available
        if 'scan_count' in globals() and scan_count > 0:
            uptime_info = f"Completed {scan_count} scans successfully"
            print_colored(f"⏱️  Session Summary: {uptime_info}", Colors.CYAN)

# ========== END OF COMPLETE MERGED SCANNER CODE ==========