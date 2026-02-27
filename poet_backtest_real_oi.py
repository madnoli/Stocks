# ==============================================================================
# ENHANCED OPTION BUYER SCANNER v3.2 - REAL OI/VOLUME ENFORCED VERSION
# No synthetic OI; all OI computations require real OpenInterest in data
# ==============================================================================

import os
import logging
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import pytz
import time
import threading
from collections import defaultdict
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import json

from tqdm import tqdm
from truedata.history import TD_hist

# Enhanced table formatting libraries
try:
    from rich.console import Console
    from rich.table import Table
    from rich.text import Text
    from rich import box
    from rich.panel import Panel
    RICH_AVAILABLE = True
except ImportError:
    RICH_AVAILABLE = False
    print("Installing rich: pip install rich")

try:
    from colorama import init, Fore, Back, Style
    init(autoreset=True)
    COLORAMA_AVAILABLE = True
except ImportError:
    COLORAMA_AVAILABLE = False
    print("Installing colorama: pip install colorama")

try:
    from great_tables import GT, md, html, style, loc
    from great_tables.data import sp500
    GREAT_TABLES_AVAILABLE = True
except ImportError:
    GREAT_TABLES_AVAILABLE = False
    print("Installing great-tables: pip install great-tables")

try:
    from tabulate import tabulate
    TABULATE_AVAILABLE = True
except ImportError:
    TABULATE_AVAILABLE = False

# Initialize console for rich output
if RICH_AVAILABLE:
    console = Console()

# Create a simple logger replacement
class Logger:
    def info(self, msg): print(f"[INFO] {msg}")
    def error(self, msg): print(f"[ERROR] {msg}")
    def warning(self, msg): print(f"[WARNING] {msg}")
    def exception(self, msg): print(f"[EXCEPTION] {msg}")

logger = Logger()

# ======== Enhanced Configuration ========
class Config:
    TDUSERNAME = os.getenv("TRUEDATA_USER", "tdwsp751")
    TDPASSWORD = os.getenv("TRUEDATA_PASS", "raj@751")

    MARKET_START = "09:15"  # IST
    FIRST_RUN_AT = "09:20"  # IST; First scan after 09:15-09:20 candle
    FIRST_SCAN_DELAY = 15   # Wait 15 seconds after 09:20 for settlement
    MARKET_END   = "15:30"  # IST
    SETTLE_DELAY_SECONDS = 15  # wait after bar close for data settlement
    MAX_WORKERS = int(os.getenv("MAX_WORKERS", "64"))
    TD_HIST_SESSIONS = int(os.getenv("TD_HIST_SESSIONS", "5"))
    SHARES_FILE = os.getenv("SHARES_FILE", "shares.txt")
    BENCHMARK_INDEX = "NIFTY 50"

    # --- Backtesting Configuration ---
    BACKTEST_INTERVAL_MINUTES = 5
    BACKTEST_START_DELAY = 5
    BACKTEST_TOP_DISPLAY = 10
    
    # --- Indicator Group Weights ---
    GROUP_WEIGHTS = {
        "Trend": 2.5, "Momentum": 2.0, "Volume": 2.2, "Volatility": 1.8, "OI": 2.5,
    }

    # --- Individual Indicator Weights within Groups ---
    INDICATOR_WEIGHTS = {
        "MA_Slope": 2.0, "ADX": 1.8, "VWAP": 1.7, "EMA": 1.7, "MACD_Trend": 1.5,
        "RSI": 2.0, "Stochastic": 1.2, "CCI": 1.2, "ROC": 1.1, "WilliamsR": 1.0,
        "VolumeSurge": 2.5, "OBV": 1.8, "CMF": 1.8, "RelVol": 1.5,
        "VolatilityExpansion": 2.5, "Bollinger": 1.3,
        "OptionBuyerMomentum": 2.8, "OIChange": 2.5, "VolumeOISync": 2.2,
    }

    # --- Scoring & Signal Thresholds ---
    SCORE_THRESHOLD_MIN = 10.0
    SIGNAL_THRESHOLDS = {
        'Very Strong Buy': 55.0, 'Strong Buy': 30.0, 'Buy Signal': 15.0,
        'Very Strong Sell': -55.0, 'Strong Sell': -30.0, 'Sell Signal': -15.0,
    }
    
    # --- Market Regime Multipliers ---
    REGIME_MULTIPLIERS = {
        'bullish_in_bull_market': 1.15, 'bearish_in_bear_market': 1.15,
        'bullish_in_bear_market': 0.8, 'bearish_in_bull_market': 0.8,
    }

# Constants
IST = pytz.timezone("Asia/Kolkata")
BAR_SIZE_MAP = {5: "5 min", 15: "15 min", 30: "30 min", 60: "60 min", 1440: "1 day"}
DURATION_MAP = {5: "30 D", 15: "30 D", 30: "60 D", 60: "120 D", 1440: "365 D"}
TIMEFRAME_WEIGHTS = {15: 3.0, 5: 2.5, 30: 2.0, 60: 1.5, 1440: 1.0}

# Silence noisy loggers
for noisy in ("truedata", "truedata.history", "truedata_ws", "websocket", "urllib3"):
    logging.getLogger(noisy).setLevel(logging.CRITICAL)

# State management
previous_scan_results = {}
previous_oi_data = {}
previous_volume_data = {}
intraday_volume_data = {}  # Track 5-minute volume changes
intraday_oi_data = {}      # Track 5-minute OI changes
scan_count = 0
backtest_stock_history = {}
current_scan_data = {}

# Color definitions
class Colors:
    HEADER = '\033[95m'; BLUE = '\033[94m'; CYAN = '\033[96m'
    GREEN = '\033[92m'; YELLOW = '\033[93m'; RED = '\033[91m'
    BOLD = '\033[1m'; UNDERLINE = '\033[4m'; END = '\033[0m'
    MAGENTA = '\033[35m'; ORANGE = '\033[33m'

def print_colored(text, color):
    if COLORAMA_AVAILABLE:
        color_map = {
            Colors.HEADER: Fore.MAGENTA + Style.BRIGHT,
            Colors.BLUE: Fore.BLUE + Style.BRIGHT,
            Colors.CYAN: Fore.CYAN + Style.BRIGHT,
            Colors.GREEN: Fore.GREEN + Style.BRIGHT,
            Colors.YELLOW: Fore.YELLOW + Style.BRIGHT,
            Colors.RED: Fore.RED + Style.BRIGHT,
            Colors.BOLD: Style.BRIGHT,
            Colors.MAGENTA: Fore.MAGENTA + Style.BRIGHT,
            Colors.ORANGE: Fore.YELLOW + Style.BRIGHT,
        }
        print(color_map.get(color, '') + text)
    else:
        print(f"{color}{text}{Colors.END}")

# ========== FIXED GREAT TABLES INTEGRATION ==========
def create_great_table_fixed(data, title, new_stocks=None, show_time=None):
    if not data or not GREAT_TABLES_AVAILABLE:
        create_rich_enhanced_table(data, title, new_stocks, show_time)
        return
    try:
        df_data = []
        for item in data:
            row = {
                'Stock': item['symbol'],
                'Signal': item['signal'],
                'Score': round(item['score'], 2),
                'Trend': round(item['sub_scores'].get('Trend', 0), 2),
                'Momentum': round(item['sub_scores'].get('Momentum', 0), 2),
                'Volume': round(item['sub_scores'].get('Volume', 0), 2),
                'OI': round(item['sub_scores'].get('OI', 0), 2),
                'Curr_Vol': item.get('current_volume', 'N/A'),
                'Curr_OI': item.get('current_oi', 'N/A'),
                'Vol_Change': item.get('volume_change_pct', 0),
                'OI_Change': item.get('oi_change_pct', 0),
                'Flow': item.get('flow', 'Unknown'),
                'Action': item.get('action', 'Consider'),
                'Is_New': 1 if (new_stocks and item['symbol'] in new_stocks) else 0
            }
            df_data.append(row)
        df = pd.DataFrame(df_data)
        gt_table = (
            GT(df)
            .tab_header(
                title=title,
                subtitle=f"Scan Time: {show_time}" if show_time else "Live Scanner Results"
            )
            .tab_spanner(label="Signal Analysis", columns=["Stock", "Signal", "Score"])
            .tab_spanner(label="Technical Indicators", columns=["Trend", "Momentum", "Volume", "OI"])
            .tab_spanner(label="Current Data", columns=["Curr_Vol", "Curr_OI", "Vol_Change", "OI_Change"])
            .tab_spanner(label="Analysis", columns=["Flow", "Action"])
            .fmt_number(columns=["Score", "Trend", "Momentum", "Volume", "OI"], decimals=2)
            .fmt_percent(columns=["Vol_Change", "OI_Change"], decimals=1)
            .data_color(columns=["Score"], palette=["red", "white", "green"], domain=[-100, 100])
            .data_color(columns=["Vol_Change"], palette=["red", "white", "lightgreen"], domain=[-50, 50])
            .data_color(columns=["OI_Change"], palette=["red", "white", "lightblue"], domain=[-50, 50])
            .tab_style(style=style.fill(color="yellow"), locations=loc.body(rows=lambda df: df['Is_New'] == 1))
            .cols_hide(columns=["Is_New"])
            .tab_options(table_font_size="12px", heading_background_color="#2E86AB", heading_text_color="white", stub_background_color="#F28E2C")
        )
        print("\n" + "="*120)
        if show_time: print(f"📊 {title} - {show_time}")
        else: print(f"📊 {title}")
        print("="*120)
        print("✨ Enhanced Great-Tables Display:")
        for i, row in df.iterrows():
            marker = "🆕 " if row['Is_New'] == 1 else "   "
            vol_chg = row['Vol_Change']
            oi_chg = row['OI_Change']
            vol_chg_str = f"{vol_chg:+.1f}%" if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1 else "N/A"
            oi_chg_str = f"{oi_chg:+.1f}%" if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1 else "N/A"
            color = Colors.MAGENTA if row['Is_New'] == 1 else Colors.END
            print_colored(f"{marker}{row['Stock']:<12} | {row['Signal']:<16} | {row['Score']:>7.2f} | {row['Curr_Vol']:>10} | {row['Curr_OI']:>10} | {vol_chg_str:>7} | {oi_chg_str:>7} | {row['Action']:<14}", color)
        print("="*120)
    except Exception as e:
        logger.error(f"Error creating great table: {e}")
        create_rich_enhanced_table(data, title, new_stocks, show_time)

def create_rich_enhanced_table(data, title, new_stocks=None, show_time=None):
    if not data:
        if RICH_AVAILABLE:
            console.print(f"\n[bold magenta]{title}[/bold magenta]")
            console.print("[yellow]No stocks found in this category.[/yellow]")
        else:
            print_colored(f"\n{title}", Colors.HEADER)
            print_colored("No stocks found in this category.", Colors.YELLOW)
        return
    if RICH_AVAILABLE:
        table = Table(box=box.ROUNDED, show_header=True, header_style="bold blue")
        table.add_column("Stock", style="bold white", width=12, justify="left")
        table.add_column("Signal", style="bold", width=16, justify="center")
        table.add_column("Score", style="bold", width=8, justify="right")
        table.add_column("Trend", style="cyan", width=7, justify="right")
        table.add_column("Momentum", style="yellow", width=9, justify="right")
        table.add_column("Volume", style="green", width=8, justify="right")
        table.add_column("OI", style="magenta", width=8, justify="right")
        table.add_column("Curr Vol", style="bright_green", width=10, justify="right")
        table.add_column("Curr OI", style="bright_magenta", width=10, justify="right")
        table.add_column("Vol Δ%", style="bright_yellow", width=8, justify="right")
        table.add_column("OI Δ%", style="bright_cyan", width=8, justify="right")
        table.add_column("Flow", style="dim", width=18, justify="left")
        table.add_column("Action", style="bold", width=14, justify="center")
        for item in data:
            symbol = item['symbol']
            is_new = new_stocks and symbol in new_stocks
            if item['score'] > 50: signal_style = "bold bright_green"
            elif item['score'] > 25: signal_style = "bold green"
            elif item['score'] > 0: signal_style = "green"
            elif item['score'] < -50: signal_style = "bold bright_red"
            elif item['score'] < -25: signal_style = "bold red"
            else: signal_style = "red"
            stock_style = "[bold bright_magenta]" + symbol + " ✨[/bold bright_magenta]" if is_new else symbol
            vol_change_raw = item.get('volume_change_pct', 0)
            if isinstance(vol_change_raw, str) or abs(vol_change_raw) < 0.1:
                vol_change_style = "[dim]N/A[/dim]"
            elif vol_change_raw > 20:
                vol_change_style = f"[bold bright_green]{vol_change_raw:+.1f}%[/bold bright_green]"
            elif vol_change_raw > 0:
                vol_change_style = f"[green]{vol_change_raw:+.1f}%[/green]"
            elif vol_change_raw < -20:
                vol_change_style = f"[bold bright_red]{vol_change_raw:+.1f}%[/bold bright_red]"
            else:
                vol_change_style = f"[red]{vol_change_raw:+.1f}%[/red]"
            oi_change_raw = item.get('oi_change_pct', 0)
            if isinstance(oi_change_raw, str) or abs(oi_change_raw) < 0.1:
                oi_change_style = "[dim]N/A[/dim]"
            elif oi_change_raw > 15:
                oi_change_style = f"[bold bright_cyan]{oi_change_raw:+.1f}%[/bold bright_cyan]"
            elif oi_change_raw > 0:
                oi_change_style = f"[cyan]{oi_change_raw:+.1f}%[/cyan]"
            elif oi_change_raw < -15:
                oi_change_style = f"[bold bright_red]{oi_change_raw:+.1f}%[/bold bright_red]"
            else:
                oi_change_style = f"[red]{oi_change_raw:+.1f}%[/red]"
            table.add_row(
                stock_style,
                f"[{signal_style}]{item['signal']}[/{signal_style}]",
                f"[bold]{item['score']:.2f}[/bold]",
                f"{item['sub_scores'].get('Trend', 0):.2f}",
                f"{item['sub_scores'].get('Momentum', 0):.2f}",
                f"{item['sub_scores'].get('Volume', 0):.2f}",
                f"{item['sub_scores'].get('OI', 0):.2f}",
                f"[bright_green]{item.get('current_volume', 'N/A')}[/bright_green]",
                f"[bright_magenta]{item.get('current_oi', 'N/A')}[/bright_magenta]",
                vol_change_style,
                oi_change_style,
                f"[dim]{item.get('flow', 'Unknown')}[/dim]",
                f"[bold]{item.get('action', 'Consider')}[/bold]"
            )
        if show_time:
            console.print(f"\n[bold magenta]{title} - {show_time}[/bold magenta]")
        else:
            console.print(f"\n[bold magenta]{title}[/bold magenta]")
        console.print(table)
    else:
        create_enhanced_ascii_table(data, title, new_stocks, show_time)

def create_enhanced_ascii_table(data, title, new_stocks=None, show_time=None):
    if not data:
        print_colored(f"\n{title}", Colors.HEADER)
        print_colored("No stocks found in this category.", Colors.YELLOW)
        return
    if show_time:
        print_colored(f"\n{title} - {show_time}", Colors.HEADER)
    else:
        print_colored(f"\n{title}", Colors.HEADER)
    print_colored("="*200, Colors.BLUE)
    header = (f"{'Stock':<12} | {'Signal':<16} | {'Score':>8} | {'Trend':>7} | {'Mom':>9} | {'Vol':>7} | {'OI':>7} | "
              f"{'CurrVol':>10} | {'CurrOI':>10} | {'VolΔ%':>8} | {'OIΔ%':>8} | {'Flow':<18} | {'Action':<14}")
    print_colored(header, Colors.BOLD)
    print_colored("-"*200, Colors.BLUE)
    for item in data:
        symbol = item['symbol']
        is_new = new_stocks and symbol in new_stocks
        vol_chg = item.get('volume_change_pct', 0)
        oi_chg = item.get('oi_change_pct', 0)
        vol_chg_str = f"{vol_chg:+.1f}" if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1 else "N/A"
        oi_chg_str = f"{oi_chg:+.1f}" if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1 else "N/A"
        row = (f"{symbol:<12} | {item['signal']:<16} | {item['score']:>8.2f} | {item['sub_scores'].get('Trend', 0):>7.2f} | "
               f"{item['sub_scores'].get('Momentum', 0):>9.2f} | {item['sub_scores'].get('Volume', 0):>7.2f} | "
               f"{item['sub_scores'].get('OI', 0):>7.2f} | {item.get('current_volume', 'N/A'):>10} | "
               f"{item.get('current_oi', 'N/A'):>10} | {vol_chg_str:>7}% | "
               f"{oi_chg_str:>7}% | {item.get('flow', 'Unknown'):<18} | {item.get('action', 'Consider'):<14}")
        if is_new:
            print_colored(row + " ← ✨ NEW!", Colors.MAGENTA)
        else:
            print(row)
    print_colored("="*200, Colors.BLUE)

def create_compact_backtest_table(data, title, new_stocks=None, show_time=None):
    if not data:
        return
    if GREAT_TABLES_AVAILABLE:
        create_great_table_fixed(data[:5], f"Compact {title}", new_stocks, show_time)
    elif RICH_AVAILABLE:
        table = Table(box=box.SIMPLE, show_header=True, header_style="bold blue")
        table.add_column("#", width=3, justify="right")
        table.add_column("Stock", style="bold white", width=12)
        table.add_column("Signal", style="bold", width=14)
        table.add_column("Score", style="bold", width=8, justify="right")
        table.add_column("CurrVol", style="bright_green", width=8, justify="right")
        table.add_column("CurrOI", style="bright_magenta", width=8, justify="right")
        table.add_column("VolΔ%", style="bright_yellow", width=7, justify="right")
        table.add_column("OIΔ%", style="bright_cyan", width=7, justify="right")
        table.add_column("Action", style="bold", width=12)
        for i, item in enumerate(data, 1):
            symbol = item['symbol']
            is_new = new_stocks and symbol in new_stocks
            stock_display = f"[bright_magenta]{symbol} ✨[/bright_magenta]" if is_new else symbol
            vol_chg = item.get('volume_change_pct', 0)
            if isinstance(vol_chg, str) or abs(vol_chg) < 0.1:
                vol_display = "[dim]N/A[/dim]"
            elif vol_chg > 0:
                vol_display = f"[bright_green]{vol_chg:+.1f}%[/bright_green]"
            else:
                vol_display = f"[red]{vol_chg:+.1f}%[/red]"
            oi_chg = item.get('oi_change_pct', 0)
            if isinstance(oi_chg, str) or abs(oi_chg) < 0.1:
                oi_display = "[dim]N/A[/dim]"
            elif oi_chg > 0:
                oi_display = f"[bright_cyan]{oi_chg:+.1f}%[/bright_cyan]"
            else:
                oi_display = f"[red]{oi_chg:+.1f}%[/red]"
            table.add_row(
                str(i),
                stock_display,
                item['signal'],
                f"{item['score']:.1f}",
                str(item.get('current_volume', 'N/A')),
                str(item.get('current_oi', 'N/A')),
                vol_display,
                oi_display,
                item.get('action', 'Consider')
            )
        if show_time:
            console.print(f"\n[bold blue]{title} - {show_time}[/bold blue]")
        else:
            console.print(f"\n[bold blue]{title}[/bold blue]")
        console.print(table)
    else:
        create_enhanced_ascii_table(data[:5], title, new_stocks, show_time)

# ========== TECHNICAL INDICATORS ==========
def ema(series, length):
    return series.ewm(span=length, adjust=False).mean()

def vwap(df, period=None):
    price = (df["High"] + df["Low"] + df["Close"]) / 3.0
    pv = price * df["Volume"]
    if period:
        pv_sum = pv.rolling(period).sum()
        vol_sum = df["Volume"].rolling(period).sum()
    else:
        pv_sum = pv.cumsum()
        vol_sum = df["Volume"].cumsum()
    return pv_sum / vol_sum.replace(0, np.nan)

def atr(df, period=14):
    high_low = df["High"] - df["Low"]
    high_close = (df["High"] - df["Close"].shift(1)).abs()
    low_close = (df["Low"] - df["Close"].shift(1)).abs()
    tr = pd.concat([high_low, high_close, low_close], axis=1).max(axis=1)
    return tr.ewm(alpha=1/period, adjust=False).mean()

def williams_r(df, period=14):
    highest = df["High"].rolling(period).max()
    lowest = df["Low"].rolling(period).min()
    return -100 * (highest - df["Close"]) / (highest - lowest).replace(0, np.nan)

def momentum(df, period=10):
    return df["Close"] / df["Close"].shift(period) - 1.0

def volume_surge(df, lookback=20):
    vol_ma = df["Volume"].rolling(lookback).mean()
    vol_std = df["Volume"].rolling(lookback).std()
    z_score = (df["Volume"] - vol_ma) / vol_std.replace(0, np.nan)
    return z_score.fillna(0)

def calculate_rsi(df, period=14):
    if len(df) < period + 1: 
        return pd.Series(dtype='float64', index=df.index)
    delta = df['Close'].diff()
    gain = (delta.where(delta > 0, 0)).ewm(com=period - 1, adjust=False).mean()
    loss = (-delta.where(delta < 0, 0)).ewm(com=period - 1, adjust=False).mean()
    rs = gain / loss.replace(0, np.nan)
    rs.fillna(100, inplace=True)
    return 100 - (100 / (1 + rs))

def calculate_macd(df, fast=12, slow=26, signal=9):
    if len(df) < slow + signal: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    ema_fast = df['Close'].ewm(span=fast, adjust=False).mean()
    ema_slow = df['Close'].ewm(span=slow, adjust=False).mean()
    macd = ema_fast - ema_slow
    signal_line = macd.ewm(span=signal, adjust=False).mean()
    return macd, signal_line

def calculate_stochastic(df, period=14, smooth_d=3):
    if len(df) < period + smooth_d: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    low_min = df['Low'].rolling(window=period).min()
    high_max = df['High'].rolling(window=period).max()
    k = 100 * ((df['Close'] - low_min) / (high_max - low_min).replace(0, np.nan))
    k.fillna(50, inplace=True)
    d = k.rolling(window=smooth_d).mean()
    return k, d
    
def calculate_adx(df, period=14):
    if len(df) < period * 2: 
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    df_adx = df.copy()
    df_adx['H-L'] = df_adx['High'] - df_adx['Low']
    df_adx['H-C'] = abs(df_adx['High'] - df_adx['Close'].shift(1))
    df_adx['L-C'] = abs(df_adx['Low'] - df_adx['Close'].shift(1))
    df_adx['TR'] = df_adx[['H-L', 'H-C', 'L-C']].max(axis=1)
    df_adx['+DM'] = np.where((df_adx['High'] - df_adx['High'].shift(1)) > (df_adx['Low'].shift(1) - df_adx['Low']), df_adx['High'] - df_adx['High'].shift(1), 0)
    df_adx['-DM'] = np.where((df_adx['Low'].shift(1) - df_adx['Low']) > (df_adx['High'] - df_adx['High'].shift(1)), df_adx['Low'].shift(1) - df_adx['Low'], 0)
    atr_val = df_adx['TR'].ewm(com=period - 1, adjust=False).mean().replace(0, np.nan)
    pdi = (df_adx['+DM'].ewm(com=period - 1, adjust=False).mean() / atr_val) * 100
    ndi = (df_adx['-DM'].ewm(com=period - 1, adjust=False).mean() / atr_val) * 100
    adx = (abs(pdi - ndi) / (pdi + ndi).replace(0, np.nan)).ewm(com=period - 1, adjust=False).mean() * 100
    return adx.fillna(20), pdi.fillna(20), ndi.fillna(20)

def calculate_bollinger_bands(df, period=20, std_dev=2):
    if len(df) < period:
        return pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index), pd.Series(dtype='float64', index=df.index)
    middle = df['Close'].rolling(window=period).mean()
    std = df['Close'].rolling(window=period).std()
    upper = middle + (std * std_dev)
    lower = middle - (std * std_dev)
    return middle, upper, lower

def calculate_roc(df, period=12):
    if len(df) < period + 1: 
        return pd.Series(dtype='float64', index=df.index)
    shifted_close = df['Close'].shift(period).replace(0, np.nan)
    return ((df['Close'] - df['Close'].shift(period)) / shifted_close) * 100

def calculate_obv(df):
    if len(df) < 2: 
        return pd.Series(dtype='float64', index=df.index)
    return (np.sign(df['Close'].diff()) * df['Volume']).fillna(0).cumsum()

def calculate_cci(df, period=20):
    if len(df) < period: 
        return pd.Series(dtype='float64', index=df.index)
    tp = (df['High'] + df['Low'] + df['Close']) / 3
    sma_tp = tp.rolling(window=period).mean()
    mad = tp.rolling(window=period).apply(lambda x: np.abs(x - x.mean()).mean(), raw=True).replace(0, np.nan)
    return (tp - sma_tp) / (0.015 * mad)
    
def cmf(df, period=20):
    mfm = ((df["Close"] - df["Low"]) - (df["High"] - df["Close"])) / (df["High"] - df["Low"]).replace(0, np.nan)
    mfm = mfm.fillna(0)
    mfv = mfm * df["Volume"]
    mfv_sum = mfv.rolling(period).sum()
    vol_sum = df["Volume"].rolling(period).sum().replace(0, np.nan)
    return (mfv_sum / vol_sum).fillna(0)
    
def relative_volume(df, lookback=50):
    vol_ma = df["Volume"].rolling(lookback).mean()
    return (df["Volume"] / vol_ma.replace(0, np.nan)).fillna(1.0)

# ---- OI metrics require real 'OpenInterest' column; otherwise return None ----
def _has_real_oi(df):
    return ('OpenInterest' in df.columns) and (df['OpenInterest'].notna().sum() >= 2)

def calculate_oi_volume_ratio(df):
    if not _has_real_oi(df):
        return pd.Series(index=df.index, dtype='float64')
    ratio = df['OpenInterest'] / df['Volume'].replace(0, np.nan)
    return ratio.fillna(0)

def detect_oi_buildup(df, lookback=20):
    if not _has_real_oi(df) or len(df) < lookback:
        return None
    oi_ma = df['OpenInterest'].rolling(lookback).mean()
    if len(oi_ma) == 0 or pd.isna(oi_ma.iloc[-1]):
        return None
    current_oi = df['OpenInterest'].iloc[-1]
    avg_oi = oi_ma.iloc[-1]
    if avg_oi > 0 and pd.notna(current_oi):
        oi_strength = (current_oi - avg_oi) / avg_oi
        return max(min(oi_strength * 100, 100), -100)
    return None

def volume_oi_sync_analysis(df):
    if len(df) < 10 or not _has_real_oi(df):
        return None
    vol_change = df['Volume'].pct_change(5).fillna(0)
    oi_change = df['OpenInterest'].pct_change(5).fillna(0)
    sync_score = vol_change.iloc[-1] + oi_change.iloc[-1]
    return min(max(sync_score * 50, -100), 100)

def option_buyer_momentum(df):
    if len(df) < 20:
        return None
    price_mom = (df['Close'].iloc[-1] / df['Close'].iloc[-5] - 1) * 100 if len(df) >= 5 else 0
    vol_surge_val = volume_surge(df, lookback=20).iloc[-1] if len(df) > 20 else 0
    oi_buildup = detect_oi_buildup(df, lookback=20)
    if oi_buildup is None:
        # If OI is unavailable, do not emit this indicator to avoid bias
        return None
    combined_score = (price_mom * 0.4) + (vol_surge_val * 0.3) + (oi_buildup * 0.3)
    return min(max(combined_score, -100), 100)

def slope(series, lookback=10):
    if len(series) < lookback: return 0.0
    y = series.tail(lookback).values
    x = np.arange(len(y))
    if len(y) < 2: return 0.0
    try:
        coeffs = np.polyfit(x, y, 1)
        return coeffs[0]
    except:
        return 0.0

# ========== SCORING ENGINE ==========
def normalize_score(value, bullish_range, bearish_range, score_range=(-2.0, 2.0)):
    low_score, high_score = score_range
    bull_min, bull_max = bullish_range
    if value >= bull_max: return high_score
    if value > bull_min:
        return high_score * ((value - bull_min) / (bull_max - bull_min))
    bear_max, bear_min = bearish_range
    if value <= bear_min: return low_score
    if value < bear_max:
        return low_score * ((bear_max - value) / (bear_max - bear_min))
    return 0.0

def calculate_indicator_scores(df):
    scores = defaultdict(float)
    if df is None or len(df) < 50: return scores
    try:
        # --- Trend Group ---
        adx, pdi, ndi = calculate_adx(df)
        if not adx.empty and len(adx) > 3 and adx.iloc[-1] > 20 and adx.iloc[-1] > adx.iloc[-3]:
            scores['ADX'] = 2.0 if pdi.iloc[-1] > ndi.iloc[-1] else -2.0
        ema20, ema50 = ema(df['Close'], 20), ema(df['Close'], 50)
        if not ema20.empty and not ema50.empty:
            ema_ratio = ema20.iloc[-1] / ema50.iloc[-1] if ema50.iloc[-1] != 0 else 1
            scores['EMA'] = normalize_score(ema_ratio, (1.001, 1.02), (0.999, 0.98))
        vwap_line = vwap(df, period=None)
        if not vwap_line.empty:
            vwap_ratio = df['Close'].iloc[-1] / vwap_line.iloc[-1] if vwap_line.iloc[-1] != 0 else 1
            scores['VWAP'] = normalize_score(vwap_ratio, (1.002, 1.025), (0.998, 0.975))
        macd, signal = calculate_macd(df)
        if not macd.empty and not signal.empty and len(macd) > 0:
            if macd.iloc[-1] > signal.iloc[-1] and macd.iloc[-1] > 0:
                scores['MACD_Trend'] = 2.0
            elif macd.iloc[-1] < signal.iloc[-1] and macd.iloc[-1] < 0:
                scores['MACD_Trend'] = -2.0
            else:
                scores['MACD_Trend'] = 0
        if not ema20.empty and len(ema20) >= 5:
            ma20_slope = slope(ema20, 5)
            price_norm_slope = ma20_slope / df['Close'].iloc[-1] * 1000 if df['Close'].iloc[-1] != 0 else 0
            scores['MA_Slope'] = normalize_score(price_norm_slope, (0.1, 0.5), (-0.1, -0.5), (-2.5, 2.5))
        
        # --- Momentum Group ---
        rsi = calculate_rsi(df)
        if not rsi.empty and len(rsi) > 0:
            scores['RSI'] = normalize_score(rsi.iloc[-1], (60, 85), (40, 15))
        k, d = calculate_stochastic(df)
        if not k.empty and not d.empty and len(k) > 0 and len(d) > 0:
            if k.iloc[-1] > d.iloc[-1]:
                scores['Stochastic'] = normalize_score(k.iloc[-1], (20, 80), (100, 100))
            elif k.iloc[-1] < d.iloc[-1]:
                scores['Stochastic'] = normalize_score(k.iloc[-1], (0,0), (80, 20))
        cci = calculate_cci(df)
        if not cci.empty and len(cci) > 0:
            scores['CCI'] = normalize_score(cci.iloc[-1], (100, 200), (-100, -200))
        roc = calculate_roc(df)
        if not roc.empty and len(roc) > 0:
            scores['ROC'] = normalize_score(roc.iloc[-1], (0.5, 2.0), (-0.5, -2.0))
        wr = williams_r(df)
        if not wr.empty and len(wr) > 0:
            scores['WilliamsR'] = normalize_score(wr.iloc[-1], (-100, -80), (-20, 0))

        # --- Volume Group ---
        zscore = volume_surge(df, lookback=20)
        if not zscore.empty and len(zscore) > 1:
            price_up = df['Close'].iloc[-1] > df['Close'].iloc[-2]
            if price_up:
                scores['VolumeSurge'] = normalize_score(zscore.iloc[-1], (1.5, 3.0), (0,0))
            else:
                scores['VolumeSurge'] = normalize_score(zscore.iloc[-1], (0,0), (-1.5, -3.0))
        obv_line = calculate_obv(df)
        if len(obv_line) > 5:
            obv_slope = slope(obv_line, 5)
            scores['OBV'] = normalize_score(obv_slope, (1, 1e9), (-1, -1e9))
        cmf20 = cmf(df, period=20)
        if not cmf20.empty and len(cmf20) > 0:
            scores['CMF'] = normalize_score(cmf20.iloc[-1], (0.1, 0.25), (-0.1, -0.25))
        rv = relative_volume(df, lookback=50)
        if not rv.empty and len(rv) > 0:
            scores['RelVol'] = normalize_score(rv.iloc[-1], (1.5, 3.0), (0.5, 0.5))

        # --- Volatility Group ---
        atr_val = atr(df, period=14)
        if len(atr_val) > 20:
            atr_ma = atr_val.rolling(20).mean()
            if len(atr_ma) > 0 and atr_ma.iloc[-1] != 0:
                atr_ratio = atr_val.iloc[-1] / atr_ma.iloc[-1]
                atr_slope_ratio = (atr_val.iloc[-1] / atr_val.iloc[-5]) if len(atr_val) >= 5 and atr_val.iloc[-5] > 0 else 1
                if atr_ratio > 1.1 and atr_slope_ratio > 1.1:
                    price_direction = 1 if df['Close'].iloc[-1] > df['Close'].iloc[-5] else -1
                    scores['VolatilityExpansion'] = 2.5 * price_direction
        _, bb_upper, bb_lower = calculate_bollinger_bands(df)
        if not bb_upper.empty and not bb_lower.empty and len(bb_upper) > 0 and len(bb_lower) > 0:
            if df['Close'].iloc[-1] > bb_upper.iloc[-1]: 
                scores['Bollinger'] = 2.0
            elif df['Close'].iloc[-1] < bb_lower.iloc[-1]: 
                scores['Bollinger'] = -2.0

        # --- OI Group (only if real OI exists) ---
        oi_bu = detect_oi_buildup(df, 20)
        if oi_bu is not None:
            scores['OIChange'] = normalize_score(oi_bu, (10, 30), (-10, -30))
        sync = volume_oi_sync_analysis(df)
        if sync is not None:
            scores['VolumeOISync'] = normalize_score(sync, (15, 40), (-15, -40))
        obm = option_buyer_momentum(df)
        if obm is not None:
            scores['OptionBuyerMomentum'] = normalize_score(obm, (20, 50), (-20, -50), (-3.0, 3.0))

    except Exception as e:
        logger.error(f"Error calculating indicator scores: {e}")
    return scores

def analyze_signals_pro(timeframe_data, market_regime='neutral'):
    total_score, total_weight = 0.0, 0.0
    group_scores = defaultdict(float)
    group_weights = defaultdict(float)
    for tf_min, df in timeframe_data.items():
        if df is None or len(df) < 50: continue
        indicator_scores = calculate_indicator_scores(df)
        tf_weight = TIMEFRAME_WEIGHTS.get(tf_min, 1.0)
        for group, weight in Config.GROUP_WEIGHTS.items():
            grp_score, grp_weight = 0.0, 0.0
            for indicator, ind_weight in Config.INDICATOR_WEIGHTS.items():
                if indicator in indicator_scores:
                    belongs_to_group = (
                        (group == 'Trend' and any(term in indicator for term in ['MA', 'ADX', 'VWAP', 'EMA', 'MACD'])) or
                        (group == 'Momentum' and any(term in indicator for term in ['RSI', 'Stochastic', 'CCI', 'ROC', 'Williams'])) or
                        (group == 'Volume' and any(term in indicator for term in ['Vol', 'OBV', 'CMF'])) or
                        (group == 'Volatility' and any(term in indicator for term in ['Volatility', 'Bollinger'])) or
                        (group == 'OI' and any(term in indicator for term in ['OI', 'Option']))
                    )
                    if belongs_to_group:
                        grp_score += indicator_scores[indicator] * ind_weight
                        grp_weight += abs(indicator_scores[indicator]) * ind_weight
            if grp_weight > 0:
                norm_grp_score = (grp_score / grp_weight) * weight * tf_weight
                group_scores[group] += norm_grp_score
                group_weights[group] += weight * tf_weight
    final_score = 0
    max_possible_score = 0
    for group, score in group_scores.items():
        final_score += score
        max_possible_score += group_weights[group]
    if max_possible_score == 0: return 'Neutral', 0.0, {}
    normalized_score = (final_score / max_possible_score) * 100
    if normalized_score > 0 and market_regime == 'bullish':
        normalized_score *= Config.REGIME_MULTIPLIERS['bullish_in_bull_market']
    elif normalized_score > 0 and market_regime == 'bearish':
        normalized_score *= Config.REGIME_MULTIPLIERS['bullish_in_bear_market']
    elif normalized_score < 0 and market_regime == 'bearish':
        normalized_score *= Config.REGIME_MULTIPLIERS['bearish_in_bear_market']
    elif normalized_score < 0 and market_regime == 'bullish':
        normalized_score *= Config.REGIME_MULTIPLIERS['bearish_in_bull_market']
    if normalized_score >= Config.SIGNAL_THRESHOLDS['Very Strong Buy']: 
        signal = 'Very Strong Buy'
    elif normalized_score >= Config.SIGNAL_THRESHOLDS['Strong Buy']: 
        signal = 'Strong Buy'
    elif normalized_score >= Config.SIGNAL_THRESHOLDS['Buy Signal']: 
        signal = 'Buy Signal'
    elif normalized_score <= Config.SIGNAL_THRESHOLDS['Very Strong Sell']: 
        signal = 'Very Strong Sell'
    elif normalized_score <= Config.SIGNAL_THRESHOLDS['Strong Sell']: 
        signal = 'Strong Sell'
    elif normalized_score <= Config.SIGNAL_THRESHOLDS['Sell Signal']: 
        signal = 'Sell Signal'
    else: 
        signal = 'Neutral'
    final_sub_scores = {}
    for group in group_scores:
        if group_weights[group] > 0:
            final_sub_scores[group] = group_scores[group] / group_weights[group] * 10
    return signal, normalized_score, final_sub_scores

# ========== 5-MINUTE VOLUME/OI TRACKING (REAL OI ONLY) ==========
def calculate_5min_volume_oi_changes(df, symbol, scan_time):
    try:
        df_5min = df[df.index <= scan_time]
        if len(df_5min) < 2:
            return 0, None, 0, 0  # current_vol, current_oi (or None), vol_change_pct, oi_change_pct
        current_volume = int(df_5min['Volume'].iloc[-1]) if 'Volume' in df_5min.columns else 0
        previous_volume = int(df_5min['Volume'].iloc[-2]) if 'Volume' in df_5min.columns else 0
        vol_change_pct = ((current_volume - previous_volume) / previous_volume * 100) if previous_volume > 0 else 0
        if _has_real_oi(df_5min):
            current_oi = int(df_5min['OpenInterest'].iloc[-1])
            previous_oi = int(df_5min['OpenInterest'].iloc[-2])
            oi_change_pct = ((current_oi - previous_oi) / previous_oi * 100) if previous_oi > 0 else 0
        else:
            current_oi, oi_change_pct = None, 0
        return current_volume, current_oi, vol_change_pct, oi_change_pct
    except Exception as e:
        logger.error(f"Error calculating 5-min changes for {symbol}: {e}")
        return 0, None, 0, 0

def extract_5min_volume_oi_data(df, symbol, time_point=None, is_live=False):
    try:
        global intraday_volume_data, intraday_oi_data

        df_slice = df[df.index <= time_point] if time_point and not is_live else df
        if df_slice.empty:
            return {
                'current_volume': 'N/A', 'current_oi': 'N/A',
                'volume_change_pct': 0, 'oi_change_pct': 0,
                'volume': 'N/A', 'oi': 'N/A', 'volume_change': 'N/A', 'oi_change': 'N/A'
            }

        current_volume, current_oi, vol_change_pct, oi_change_pct = calculate_5min_volume_oi_changes(
            df_slice, symbol, df_slice.index[-1]
        )

        if abs(vol_change_pct) < 0.1 and abs(oi_change_pct) < 0.1:
            prev_volume = intraday_volume_data.get(symbol, None)
            prev_oi = intraday_oi_data.get(symbol, None)
            if prev_volume is not None and prev_volume > 0 and current_volume and current_volume > 0:
                vol_change_pct = ((current_volume - prev_volume) / prev_volume) * 100
            if prev_oi is not None and prev_oi > 0 and current_oi and current_oi > 0:
                oi_change_pct = ((current_oi - prev_oi) / prev_oi) * 100

        intraday_volume_data[symbol] = current_volume if isinstance(current_volume, int) else 0
        intraday_oi_data[symbol] = current_oi if isinstance(current_oi, int) else 0

        current_volume_display = f"{current_volume:,}" if isinstance(current_volume, int) and current_volume > 999 else str(current_volume)
        current_oi_display = f"{current_oi:,}" if isinstance(current_oi, int) and current_oi and current_oi > 999 else (str(current_oi) if current_oi is not None else "N/A")

        volume_change_legacy = f"{vol_change_pct:+.1f}%" if isinstance(vol_change_pct, (int, float)) and abs(vol_change_pct) > 0.1 else "N/A"
        oi_change_legacy = f"{oi_change_pct:+.1f}%" if isinstance(oi_change_pct, (int, float)) and abs(oi_change_pct) > 0.1 else "N/A"

        return {
            'current_volume': current_volume_display,
            'current_oi': current_oi_display,
            'volume_change_pct': vol_change_pct if isinstance(vol_change_pct, (int, float)) and abs(vol_change_pct) > 0.1 else 0,
            'oi_change_pct': oi_change_pct if isinstance(oi_change_pct, (int, float)) and abs(oi_change_pct) > 0.1 else 0,
            'volume': current_volume_display,
            'oi': current_oi_display,
            'volume_change': volume_change_legacy,
            'oi_change': oi_change_legacy,
            '_raw_volume': current_volume if isinstance(current_volume, int) else 0,
            '_raw_oi': current_oi if isinstance(current_oi, int) else 0
        }
    except Exception as e:
        logger.error(f"Error extracting 5-min data for {symbol}: {e}")
        return {
            'current_volume': 'N/A', 'current_oi': 'N/A',
            'volume_change_pct': 0, 'oi_change_pct': 0,
            'volume': 'N/A', 'oi': 'N/A', 'volume_change': 'N/A', 'oi_change': 'N/A'
        }

# ========== TIMING FUNCTIONS ==========
def generate_backtest_timestamps(backtest_date):
    timestamps = []
    base_date = IST.localize(datetime.strptime(backtest_date, "%Y-%m-%d"))
    current_time = base_date.replace(hour=9, minute=15, second=0, microsecond=0)
    market_end = base_date.replace(hour=15, minute=30, second=0, microsecond=0)
    first_scan = current_time + timedelta(minutes=5, seconds=Config.SETTLE_DELAY_SECONDS)
    timestamps.append(first_scan)
    current_scan = first_scan
    while current_scan < market_end:
        current_scan += timedelta(minutes=5)
        if current_scan <= market_end:
            timestamps.append(current_scan)
    return timestamps

def next_5min_boundary_ist(now_ist: datetime) -> datetime:
    minute = (now_ist.minute // 5) * 5
    boundary = now_ist.replace(minute=minute, second=0, microsecond=0)
    if boundary <= now_ist:
        boundary = boundary + timedelta(minutes=5)
    return boundary

def get_exact_candle_close_time(now_ist: datetime) -> datetime:
    next_boundary = next_5min_boundary_ist(now_ist)
    return next_boundary + timedelta(seconds=Config.SETTLE_DELAY_SECONDS)

def parse_hhmm(s: str):
    h, m = map(int, s.split(":"))
    return h, m

def today_ist_dt(hhmm: str) -> datetime:
    now = datetime.now(IST)
    h, m = parse_hhmm(hhmm)
    return now.replace(hour=h, minute=m, second=0, microsecond=0)

def sleep_until(ts: datetime):
    while True:
        now = datetime.now(IST)
        delta = (ts - now).total_seconds()
        if delta <= 0:
            break
        time.sleep(min(0.5, delta))

# ========== DATA FETCHING ==========
class TokenBucketLimiter:
    def __init__(self, rate_per_sec: float, bucket_size: int):
        self.rate = rate_per_sec
        self.capacity = bucket_size
        self.tokens = bucket_size
        self.lock = threading.Lock()
        self.last_refill = time.time()
    def acquire(self):
        while True:
            with self.lock:
                now = time.time()
                elapsed = now - self.last_refill
                if elapsed > 0:
                    add = int(elapsed * self.rate)
                    if add > 0:
                        self.tokens = min(self.capacity, self.tokens + add)
                        self.last_refill = now
                if self.tokens > 0:
                    self.tokens -= 1
                    return
                sleep_for = max(0.0, 1.0 / self.rate)
            time.sleep(sleep_for)

api_calls_done = 0
api_calls_lock = threading.Lock()

def authenticate_session():
    return TD_hist(Config.TDUSERNAME, Config.TDPASSWORD, log_level=logging.CRITICAL)

def build_sessions():
    pool = []
    for i in range(Config.TD_HIST_SESSIONS):
        try:
            pool.append(authenticate_session())
        except Exception as e:
            logger.error(f"Session {i} init failed: {e}")
    if not pool:
        raise SystemExit("Failed to initialize TrueData sessions.")
    per_sess_rate = 10.0 / len(pool)
    limiters = [TokenBucketLimiter(rate_per_sec=per_sess_rate, bucket_size=10) for _ in pool]
    return pool, limiters

tdhist_pool, sess_limiters = build_sessions()
logger.info("TrueData login successful.")

def normalize_hist_df(df, symbol):
    if df is None or len(df) == 0: return None
    try:
        out = df.copy()
        out.rename(columns={c: str(c).lower() for c in out.columns}, inplace=True)
        rename_map = {}
        for src, tgt in (("timestamp","Date"),("time","Date"),("datetime","Date"),("date","Date"),
                         ("open","Open"),("high","High"),("low","Low"),("close","Close"),
                         ("volume","Volume"),("vol","Volume"),
                         ("oi","OpenInterest"),("openinterest","OpenInterest"),("open_interest","OpenInterest")):
            if src in out.columns: rename_map[src] = tgt
        out.rename(columns=rename_map, inplace=True)
        if "Date" not in out.columns and isinstance(out.index, pd.DatetimeIndex):
            out["Date"] = out.index
        elif "Date" not in out.columns:
            return None
        if "Volume" not in out.columns:
            out["Volume"] = 0
        if "OpenInterest" in out.columns:
            out["OpenInterest"] = pd.to_numeric(out["OpenInterest"], errors="coerce")
        out["Date"] = pd.to_datetime(out["Date"], errors="coerce")
        out = out.dropna(subset=["Date"])
        if pd.api.types.is_datetime64tz_dtype(out["Date"]):
            out["Date"] = out["Date"].dt.tz_convert(IST)
        else:
            out["Date"] = out["Date"].dt.tz_localize(IST)
        for c in ["Open","High","Low","Close","Volume"]:
            out[c] = pd.to_numeric(out.get(c, np.nan), errors="coerce")
        out = out.dropna(subset=["Open","High","Low","Close"]).sort_values("Date").set_index("Date")
        out = out[~out.index.duplicated(keep='last')]
        return out
    except Exception as e:
        logger.error(f"Normalize error {symbol}: {e}")
        return None

def pick_session(symbol_orig, timeframe_minutes):
    return (hash(symbol_orig) ^ timeframe_minutes) % len(tdhist_pool)

def fetch_one_timeaware(symbol_orig, timeframe_minutes, limiter, hist, up_to_time):
    td_symbol = symbol_orig.replace('-EQ', '')
    bar_size, duration = BAR_SIZE_MAP.get(timeframe_minutes), DURATION_MAP.get(timeframe_minutes)
    if not bar_size or not duration: return symbol_orig, timeframe_minutes, None
    try:
        limiter.acquire()
        df_raw = hist.get_historic_data(td_symbol, duration=duration, bar_size=bar_size)
        df = normalize_hist_df(df_raw, td_symbol)
        if df is not None and up_to_time:
             if isinstance(df.index, pd.DatetimeIndex) and not df.index.isnull().any():
                  try:
                       df = df[df.index <= up_to_time]
                  except (TypeError, ValueError) as filter_err:
                        logger.error(f"Filtering failed for {symbol_orig} {timeframe_minutes}min: {filter_err}")
                        return symbol_orig, timeframe_minutes, None
        else:
            logger.warning(f"Invalid or incomplete datetime index for {symbol_orig} {timeframe_minutes}min, skipping filter")
                return symbol_orig, timeframe_minutes, None
        global api_calls_done
        with api_calls_lock: api_calls_done += 1
        return symbol_orig, timeframe_minutes, df
    except Exception as e:
        logger.error(f"Error fetching {symbol_orig} {timeframe_minutes}min: {e}")
        return symbol_orig, timeframe_minutes, None

def fetch_one(symbol_orig, timeframe_minutes, limiter, hist):
    return fetch_one_timeaware(symbol_orig, timeframe_minutes, limiter, hist, None)

def prefetch_all_timeaware(stocks, up_to_time=None, max_workers=Config.MAX_WORKERS):
    tfs = [5, 15, 30, 60, 1440]
    total_calls, stock_multi_data = len(stocks) * len(tfs), defaultdict(dict)
    global api_calls_done
    with api_calls_lock: api_calls_done = 0
    desc = f"Fetching data (up to {up_to_time.strftime('%H:%M')})" if up_to_time else "Prefetching Data"
    with tqdm(total=total_calls, desc=desc, ncols=100, leave=False) as api_bar:
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = []
            for s in stocks:
                for tf in tfs:
                    session_idx = pick_session(s, tf)
                    if up_to_time:
                        futures.append(executor.submit(
                            fetch_one_timeaware, s, tf, sess_limiters[session_idx],
                            tdhist_pool[session_idx], up_to_time
                        ))
                    else:
                        futures.append(executor.submit(
                            fetch_one, s, tf, sess_limiters[session_idx], tdhist_pool[session_idx]
                        ))
            for fut in as_completed(futures):
                symbol_orig, tf, df = fut.result()
                if df is not None and len(df) > 0:
                    stock_multi_data[symbol_orig][tf] = df
                api_bar.update(1)
    return {s: d for s, d in stock_multi_data.items() if len(d) >= 2}

def prefetch_all(stocks, max_workers=Config.MAX_WORKERS):
    return prefetch_all_timeaware(stocks, None, max_workers)

def get_market_regime(index_symbol="NIFTY 50"):
    try:
        si = pick_session(index_symbol, 1440)
        df_raw = tdhist_pool[si].get_historic_data(index_symbol, duration="200 D", bar_size="1 day")
        df = normalize_hist_df(df_raw, index_symbol)
        if df is None or len(df) < 50: return 'neutral'
        ema20_series = ema(df['Close'], 20)
        ema50_series = ema(df['Close'], 50)
        if ema20_series.empty or ema50_series.empty: return 'neutral'
        ema20_val = ema20_series.iloc[-1]
        ema50_val = ema50_series.iloc[-1]
        close = df['Close'].iloc[-1]
        if close > ema20_val and ema20_val > ema50_val:
            return 'bullish'
        elif close < ema20_val and ema20_val < ema50_val:
            return 'bearish'
        else:
            return 'neutral'
    except Exception as e:
        logger.warning(f"Could not fetch market regime for {index_symbol}: {e}")
        return 'neutral'

def enhanced_institutional_flow_analysis(tf_data):
    frames = [tf_data.get(t) for t in (5, 15, 30) if tf_data.get(t) is not None and len(tf_data.get(t)) >= 60]
    if not frames: return "Unknown"
    votes = 0
    for df in frames:
        cmf_series = cmf(df, 20)
        rv_series = relative_volume(df, 50)
        if cmf_series.empty or rv_series.empty: continue
        cmf_last = cmf_series.iloc[-1]
        rv_last = rv_series.iloc[-1]
        if cmf_last > 0.05 and rv_last > 1.2: votes += 1
        elif cmf_last < -0.05 and rv_last > 1.2: votes -= 1
    if votes >= 2: return "Institutional Accumulation"
    elif votes <= -2: return "Institutional Distribution"
    else: return "Mixed/Neutral"

# ========== MAIN SCANNER LOGIC ==========
def run_scan_at_time_5min_fixed(time_point_aware, stocks, market_regime, is_live=False):
    stock_multi_data = prefetch_all(stocks, max_workers=Config.MAX_WORKERS) if is_live else \
                       prefetch_all_timeaware(stocks, time_point_aware, max_workers=Config.MAX_WORKERS)
    print_colored(f"Data fetch complete. Analyzing signals (Market Regime: {market_regime.upper()})...", Colors.CYAN)
    signals_this_scan = []
    current_symbols = set()
    for symbol, timeframe_data in stock_multi_data.items():
        clean_symbol = symbol.replace('-EQ', '')
        current_symbols.add(clean_symbol)
        filtered_timeframes = {}
        for tf, df in timeframe_data.items():
            if df is not None and not df.empty:
                df_slice = df if is_live else df[df.index <= time_point_aware]
                if not df_slice.empty and len(df_slice) >= 50:
                    filtered_timeframes[tf] = df_slice
        if len(filtered_timeframes) < 2: continue
        signal, score, sub_scores = analyze_signals_pro(filtered_timeframes, market_regime)
        if abs(score) >= Config.SCORE_THRESHOLD_MIN:
            flow_tag = enhanced_institutional_flow_analysis(filtered_timeframes)
            tf_5min = filtered_timeframes.get(5)
            if tf_5min is not None:
                oi_vol_data = extract_5min_volume_oi_data(tf_5min, clean_symbol, time_point_aware, is_live=is_live)
            else:
                main_tf_data = filtered_timeframes.get(15, filtered_timeframes.get(30, list(filtered_timeframes.values())[0]))
                oi_vol_data = extract_5min_volume_oi_data(main_tf_data, clean_symbol, time_point_aware, is_live=is_live)
            action = "Consider Call" if score > 0 else "Consider Put"
            if 'Strong' in signal: action = f"Strong {'Call' if score > 0 else 'Put'} Buy"
            signals_this_scan.append({
                'symbol': clean_symbol,
                'signal': signal,
                'score': score,
                'sub_scores': sub_scores,
                'flow': flow_tag,
                'action': action,
                **oi_vol_data
            })
    return signals_this_scan, current_symbols

# (Backtest, export, report, and main() remain as in the prior block)
# ========== FIXED BACKTEST FUNCTION ==========
def run_full_day_backtest_5min_fixed(backtest_date, stocks):
    """FIXED: Enhanced backtest with proper 5-minute Volume/OI change tracking."""
    global backtest_stock_history, intraday_volume_data, intraday_oi_data
    
    print_colored(f"\n🔄 STARTING FIXED 5-MIN BACKTEST FOR {backtest_date}", Colors.HEADER)
    
    if GREAT_TABLES_AVAILABLE:
        print_colored("✅ Using Fixed Great-Tables for enhanced visualization", Colors.CYAN)
    
    timestamps = generate_backtest_timestamps(backtest_date)
    total_scans = len(timestamps)
    
    print_colored(f"Generated {total_scans} scan points from {timestamps[0].strftime('%H:%M')} to {timestamps[-1].strftime('%H:%M')}", Colors.CYAN)
    
    market_regime = get_market_regime(Config.BENCHMARK_INDEX)
    print_colored(f"Market Regime: {market_regime.upper()}", Colors.BLUE)
    
    all_results = []
    backtest_stock_history = {}
    intraday_volume_data = {}  # Reset for backtest
    intraday_oi_data = {}      # Reset for backtest
    
    with tqdm(total=total_scans, desc="Fixed 5-Min Backtesting", ncols=120) as pbar:
        for i, scan_time in enumerate(timestamps):
            try:
                pbar.set_description(f"Scanning at {scan_time.strftime('%H:%M:%S')}")
                
                # FIXED: Use the 5-minute corrected scan function
                signals, current_symbols = run_scan_at_time_5min_fixed(scan_time, stocks, market_regime, is_live=False)
                
                previous_symbols = set(backtest_stock_history.keys())
                new_stocks = current_symbols - previous_symbols
                
                for symbol in current_symbols:
                    backtest_stock_history[symbol] = scan_time
                
                scan_result = {
                    'timestamp': scan_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'scan_number': i + 1,
                    'total_signals': len(signals),
                    'bullish_signals': len([s for s in signals if s['score'] > 0]),
                    'bearish_signals': len([s for s in signals if s['score'] < 0]),
                    'new_stocks': list(new_stocks),
                    'signals': signals
                }
                all_results.append(scan_result)
                
                if signals:
                    signals.sort(key=lambda x: abs(x['score']), reverse=True)
                    top_bullish = [r for r in signals if r['score'] > 0][:Config.BACKTEST_TOP_DISPLAY]
                    top_bearish = [r for r in signals if r['score'] < 0][:Config.BACKTEST_TOP_DISPLAY]
                    
                    scan_time_str = scan_time.strftime('%H:%M')
                    
                    # Count stocks with meaningful volume/OI changes
                    vol_with_changes = sum(1 for s in signals if isinstance(s.get('volume_change_pct', 0), (int, float)) and abs(s.get('volume_change_pct', 0)) > 0.1)
                    oi_with_changes = sum(1 for s in signals if isinstance(s.get('oi_change_pct', 0), (int, float)) and abs(s.get('oi_change_pct', 0)) > 0.1)
                    
                    if RICH_AVAILABLE:
                        console.print(f"\n[bold blue]🔥 SCAN #{i+1}/{total_scans} - {scan_time_str} IST[/bold blue]")
                        console.print(f"[cyan]Signals: {len(signals)} | Bullish: {len([s for s in signals if s['score'] > 0])} | Bearish: {len([s for s in signals if s['score'] < 0])} | New: {len(new_stocks)}[/cyan]")
                        console.print(f"[yellow]Volume Changes: {vol_with_changes} stocks | OI Changes: {oi_with_changes} stocks[/yellow]")
                    else:
                        print_colored(f"\n🔥 SCAN #{i+1}/{total_scans} - {scan_time_str} IST", Colors.BOLD)
                        print_colored(f"Signals: {len(signals)} | Bullish: {len([s for s in signals if s['score'] > 0])} | Bearish: {len([s for s in signals if s['score'] < 0])} | New: {len(new_stocks)}", Colors.CYAN)
                        print_colored(f"Volume Changes: {vol_with_changes} stocks | OI Changes: {oi_with_changes} stocks", Colors.YELLOW)
                    
                    if top_bullish:
                        # Use fixed great table function
                        if GREAT_TABLES_AVAILABLE:
                            create_great_table_fixed(top_bullish, "🟢 TOP BULLISH", new_stocks, scan_time_str)
                        else:
                            create_compact_backtest_table(top_bullish, "🟢 TOP BULLISH", new_stocks, scan_time_str)
                    
                    if top_bearish:
                        # Use fixed great table function
                        if GREAT_TABLES_AVAILABLE:
                            create_great_table_fixed(top_bearish, "🔴 TOP BEARISH", new_stocks, scan_time_str)
                        else:
                            create_compact_backtest_table(top_bearish, "🔴 TOP BEARISH", new_stocks, scan_time_str)
                    
                    if new_stocks and len(new_stocks) > 0:
                        new_stocks_display = list(new_stocks)[:10]
                        more_text = f"... +{len(new_stocks)-10}" if len(new_stocks) > 10 else ""
                        print_colored(f"\n✨ NEW STOCKS: {', '.join(new_stocks_display)}{more_text}", Colors.MAGENTA)
                        
                else:
                    print_colored(f"\n[{scan_time.strftime('%H:%M')}] Scan #{i+1}/{total_scans} - No signals", Colors.YELLOW)
                
                pbar.update(1)
                time.sleep(0.1)
                    
            except Exception as e:
                logger.error(f"Error in backtest scan at {scan_time}: {e}")
                pbar.update(1)
                continue
    
    # Show enhanced summary
    print_colored(f"\n📊 FIXED 5-MIN BACKTEST SUMMARY FOR {backtest_date}", Colors.HEADER)
    print_colored("="*120, Colors.BLUE)
    
    total_scans_completed = len([r for r in all_results if r['total_signals'] >= 0])
    total_signals = sum(r['total_signals'] for r in all_results)
    total_bullish = sum(r['bullish_signals'] for r in all_results)
    total_bearish = sum(r['bearish_signals'] for r in all_results)
    unique_stocks = len(backtest_stock_history)
    
    print(f"📈 Scans: {total_scans_completed}/{total_scans}")
    print(f"📊 Total Signals: {total_signals}")
    print(f"🟢 Bullish: {total_bullish}")
    print(f"🔴 Bearish: {total_bearish}")
    print(f"📋 Unique Stocks: {unique_stocks}")
    
    if total_signals > 0:
        print(f"📊 Avg Signals/Scan: {total_signals/total_scans_completed:.1f}")
        print(f"⚖️  Bull/Bear Ratio: {total_bullish/max(total_bearish, 1):.2f}")
    
    # FIXED: Enhanced Volume/OI change statistics
    print_colored("\n📊 5-MINUTE VOLUME/OI CHANGE STATISTICS:", Colors.CYAN)
    vol_changes = []
    oi_changes = []
    
    for result in all_results[1:]:  # Skip first scan
        for signal in result['signals']:
            vol_chg = signal.get('volume_change_pct', 0)
            oi_chg = signal.get('oi_change_pct', 0)
            if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 0.1:
                vol_changes.append(vol_chg)
            if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 0.1:
                oi_changes.append(oi_chg)
    
    if vol_changes:
        avg_vol_chg = sum(vol_changes) / len(vol_changes)
        max_vol_chg = max(vol_changes)
        min_vol_chg = min(vol_changes)
        print(f"  📈 Volume Changes: {len(vol_changes)} stocks with meaningful changes")
        print(f"     Average: {avg_vol_chg:.1f}% | Max: {max_vol_chg:.1f}% | Min: {min_vol_chg:.1f}%")
    else:
        print(f"  📈 Volume Changes: No meaningful changes detected (threshold: >0.1%)")
    
    if oi_changes:
        avg_oi_chg = sum(oi_changes) / len(oi_changes)
        max_oi_chg = max(oi_changes)
        min_oi_chg = min(oi_changes)
        print(f"  📊 OI Changes: {len(oi_changes)} stocks with meaningful changes")
        print(f"     Average: {avg_oi_chg:.1f}% | Max: {max_oi_chg:.1f}% | Min: {min_oi_chg:.1f}%")
    else:
        print(f"  📊 OI Changes: No meaningful changes detected (threshold: >0.1%)")
    
    # Most active times
    active_scans = sorted(all_results, key=lambda x: x['total_signals'], reverse=True)[:5]
    print_colored("\n🔥 MOST ACTIVE TIMES:", Colors.CYAN)
    for i, scan in enumerate(active_scans):
        if scan['total_signals'] > 0:
            time_str = datetime.fromisoformat(scan['timestamp']).strftime('%H:%M')
            vol_active = sum(1 for s in scan['signals'] if isinstance(s.get('volume_change_pct', 0), (int, float)) and abs(s.get('volume_change_pct', 0)) > 0.1)
            print(f"  {i+1}. {time_str} - {scan['total_signals']} signals ({scan['bullish_signals']}B/{scan['bearish_signals']}S) | {vol_active} vol changes")
    
    # Save results
    output_filename = f"{backtest_date}_5min_fixed_backtest_results.json"
    try:
        with open(output_filename, 'w') as f:
            json.dump(all_results, f, indent=2)
        print_colored(f"\n💾 Results saved: {output_filename}", Colors.GREEN)
    except Exception as e:
        logger.error(f"Could not save results: {e}")
    
    print_colored("="*120, Colors.BLUE)
    print_colored("🎯 Fixed 5-minute backtesting completed!", Colors.GREEN)

# ========== ADDITIONAL UTILITY FUNCTIONS ==========
def export_backtest_to_excel(results, filename):
    """Export backtest results to Excel with multiple sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add summary data
        summary_data = [
            ["Metric", "Value"],
            ["Total Scans", len(results)],
            ["Total Signals", sum(r['total_signals'] for r in results)],
            ["Total Bullish", sum(r['bullish_signals'] for r in results)],
            ["Total Bearish", sum(r['bearish_signals'] for r in results)],
            ["Unique Stocks", len(set(stock for r in results for stock in r.get('new_stocks', [])))],
        ]
        
        for row_idx, (metric, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=metric)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Signals sheet
        ws_signals = wb.create_sheet("All Signals")
        headers = ["Timestamp", "Stock", "Signal", "Score", "Action", "Volume Change", "OI Change"]
        
        for col_idx, header in enumerate(headers, 1):
            ws_signals.cell(row=1, column=col_idx, value=header)
        
        row_idx = 2
        for result in results:
            for signal in result.get('signals', []):
                ws_signals.cell(row=row_idx, column=1, value=result['timestamp'])
                ws_signals.cell(row=row_idx, column=2, value=signal['symbol'])
                ws_signals.cell(row=row_idx, column=3, value=signal['signal'])
                ws_signals.cell(row=row_idx, column=4, value=signal['score'])
                ws_signals.cell(row=row_idx, column=5, value=signal.get('action', 'N/A'))
                ws_signals.cell(row=row_idx, column=6, value=signal.get('volume_change', 'N/A'))
                ws_signals.cell(row=row_idx, column=7, value=signal.get('oi_change', 'N/A'))
                row_idx += 1
        
        wb.save(filename)
        print_colored(f"📊 Excel export saved: {filename}", Colors.GREEN)
        
    except ImportError:
        print_colored("⚠️  openpyxl not installed. Install with: pip install openpyxl", Colors.YELLOW)
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}")

def generate_performance_report(results, backtest_date):
    """Generate a comprehensive performance report."""
    print_colored(f"\n📈 COMPREHENSIVE PERFORMANCE REPORT - {backtest_date}", Colors.HEADER)
    print_colored("="*100, Colors.BLUE)
    
    if not results:
        print_colored("No data available for report generation.", Colors.YELLOW)
        return
    
    # Time-based analysis
    hourly_signals = defaultdict(int)
    for result in results:
        timestamp = datetime.fromisoformat(result['timestamp'])
        hour = timestamp.hour
        hourly_signals[hour] += result['total_signals']
    
    print_colored("\n⏰ HOURLY ACTIVITY BREAKDOWN:", Colors.CYAN)
    for hour in sorted(hourly_signals.keys()):
        bar_length = min(50, int(hourly_signals[hour] * 50 / max(hourly_signals.values())))
        bar = "█" * bar_length
        print(f"  {hour:02d}:00 | {hourly_signals[hour]:4d} signals | {bar}")
    
    # Signal strength distribution
    strong_signals = sum(len([s for s in r['signals'] if 'Strong' in s.get('signal', '')]) for r in results)
    total_signals = sum(r['total_signals'] for r in results)
    
    print_colored(f"\n💪 SIGNAL STRENGTH ANALYSIS:", Colors.CYAN)
    print(f"  Strong Signals: {strong_signals}/{total_signals} ({strong_signals/max(total_signals,1)*100:.1f}%)")
    
    # Volume/OI change analysis
    significant_vol_changes = 0
    significant_oi_changes = 0
    
    for result in results:
        for signal in result.get('signals', []):
            vol_chg = signal.get('volume_change_pct', 0)
            oi_chg = signal.get('oi_change_pct', 0)
            
            if isinstance(vol_chg, (int, float)) and abs(vol_chg) > 20:
                significant_vol_changes += 1
            if isinstance(oi_chg, (int, float)) and abs(oi_chg) > 15:
                significant_oi_changes += 1
    
    print_colored(f"\n📊 SIGNIFICANT CHANGES:", Colors.CYAN)
    print(f"  Volume Changes >20%: {significant_vol_changes}")
    print(f"  OI Changes >15%: {significant_oi_changes}")
    
    # Export option
    export_filename = f"{backtest_date}_performance_report.xlsx"
    try:
        export_backtest_to_excel(results, export_filename)
    except Exception as e:
        logger.error(f"Could not create Excel export: {e}")
    
    print_colored("="*100, Colors.BLUE)

# ========== MAIN FUNCTION ==========
def main_final_fixed():
    parser = argparse.ArgumentParser(description="FINAL FIXED Enhanced Options Buyer Scanner v3.2")
    parser.add_argument("--asof", type=str, help="Backtest snapshot: 2025-09-30T14:50")
    parser.add_argument("--backtest", type=str, help="Full day backtest: 2025-09-30")
    args = parser.parse_args()

    try:
        with open(Config.SHARES_FILE, 'r') as f:
            stocks = [line.strip().upper() for line in f.readlines() if line.strip()]
        logger.info(f"Loaded {len(stocks)} symbols from {Config.SHARES_FILE}")
    except Exception:
        stocks = ["RELIANCE-EQ", "TCS-EQ", "HDFCBANK-EQ", "INFY-EQ", "HINDUNILVR-EQ", "ICICIBANK-EQ", "SBIN-EQ"]
        logger.warning(f"Could not load {Config.SHARES_FILE}. Using sample stocks.")

    if args.backtest:
        try:
            datetime.strptime(args.backtest, "%Y-%m-%d")
            run_full_day_backtest_5min_fixed(args.backtest, stocks)  # FINAL FIXED function
        except ValueError:
            logger.error("Invalid date format for --backtest. Use YYYY-MM-DD.")
            return
    elif args.asof:
        # Use fixed snapshot function
        try:
            asof_ts = IST.localize(datetime.fromisoformat(args.asof))
        except ValueError:
            try:
                asof_ts = IST.localize(datetime.strptime(args.asof, "%Y-%m-%d"))
                asof_ts = asof_ts.replace(hour=15, minute=30)
            except ValueError:
                logger.error(f"Invalid timestamp format: {args.asof}")
                return
        
        logger.info(f"Running final fixed snapshot for: {asof_ts.strftime('%Y-%m-%d %H:%M:%S %Z')}")
        market_regime = get_market_regime(Config.BENCHMARK_INDEX)
        signals, _ = run_scan_at_time_5min_fixed(asof_ts, stocks, market_regime, is_live=False)

        signals.sort(key=lambda x: abs(x['score']), reverse=True)
        top_bullish = [r for r in signals if r['score'] > 0][:10]
        top_bearish = [r for r in signals if r['score'] < 0][:10]
        
        print_colored(f"\nFINAL FIXED SNAPSHOT RESULTS - {asof_ts.strftime('%Y-%m-%d %H:%M')} IST", Colors.BOLD)
        
        if GREAT_TABLES_AVAILABLE:
            create_great_table_fixed(top_bullish, "🟢 TOP 10 BULLISH OPPORTUNITIES")
            create_great_table_fixed(top_bearish, "🔴 TOP 10 BEARISH OPPORTUNITIES")
        else:
            create_rich_enhanced_table(top_bullish, "🟢 TOP 10 BULLISH OPPORTUNITIES")
            create_rich_enhanced_table(top_bearish, "🔴 TOP 10 BEARISH OPPORTUNITIES")
    else:
        # Live scanner with fixed functions
        print_colored("\n🚀 STARTING FINAL FIXED LIVE SCANNER v3.2", Colors.GREEN)
        print_colored("✅ Fixed: Great Tables + 5-Min Volume/OI Tracking", Colors.CYAN)
        
        # Use the existing live scanner but with fixed scan function
        global scan_count, previous_scan_results, intraday_volume_data, intraday_oi_data
        
        try:
            with open(Config.SHARES_FILE, 'r') as f:
                stocks = [line.strip().upper() for line in f.readlines() if line.strip()]
            logger.info(f"Loaded {len(stocks)} symbols from {Config.SHARES_FILE}")
        except Exception as e:
            raise SystemExit(f"Could not read {Config.SHARES_FILE}: {e}")

        # Initialize and run live scanner with fixed functions
        intraday_volume_data = {}
        intraday_oi_data = {}
        scan_count = 0
        previous_scan_results = {}
        
        # Wait for first scan at 09:20:15
        now_ist = datetime.now(IST)
        first_run_time = today_ist_dt(Config.FIRST_RUN_AT)
        first_scan_time = first_run_time + timedelta(seconds=Config.FIRST_SCAN_DELAY)
        
        if now_ist < first_scan_time:
            logger.info(f"Waiting until {first_scan_time.strftime('%H:%M:%S')} IST for first scan...")
            sleep_until(first_scan_time)
        
        while True:
            scan_count += 1
            now_ist = datetime.now(IST)
            
            if now_ist.time() > datetime.strptime(Config.MARKET_END, "%H:%M").time():
                logger.info("Market closed. Shutting down.")
                break

            print_colored(f"\n[{now_ist.strftime('%H:%M:%S')}] FINAL FIXED SCANNER v3.2 - Scan #{scan_count}", Colors.HEADER)
            
            market_regime = get_market_regime(Config.BENCHMARK_INDEX)
            signals, current_symbols = run_scan_at_time_5min_fixed(now_ist, stocks, market_regime, is_live=True)
            
            new_stocks = current_symbols - set(previous_scan_results.keys()) if previous_scan_results else set()
            previous_scan_results = {s: True for s in current_symbols}
            
            signals.sort(key=lambda x: abs(x['score']), reverse=True)
            top_bullish = [r for r in signals if r['score'] > 0][:10]
            top_bearish = [r for r in signals if r['score'] < 0][:10]

            print_colored(f"\nFINAL FIXED SCANNER RESULTS - {now_ist.strftime('%Y-%m-%d %H:%M')} IST (Regime: {market_regime.upper()})", Colors.BOLD)
            
            if GREAT_TABLES_AVAILABLE:
                create_great_table_fixed(top_bullish, "🟢 TOP 10 BULLISH OPPORTUNITIES", new_stocks)
                create_great_table_fixed(top_bearish, "🔴 TOP 10 BEARISH OPPORTUNITIES", new_stocks)
            else:
                create_rich_enhanced_table(top_bullish, "🟢 TOP 10 BULLISH OPPORTUNITIES", new_stocks)
                create_rich_enhanced_table(top_bearish, "🔴 TOP 10 BEARISH OPPORTUNITIES", new_stocks)
            
            next_scan_time = get_exact_candle_close_time(datetime.now(IST))
            logger.info(f"Next scan at {next_scan_time.strftime('%H:%M:%S')}.")
            sleep_until(next_scan_time)

# ========== MAIN EXECUTION WITH ENHANCED ERROR HANDLING ==========
if __name__ == "__main__":
    try:
        # Display startup banner
        print_colored("\n🎯 FINAL FIXED ENHANCED OPTION BUYER SCANNER v3.2", Colors.HEADER)
        print_colored("✅ Fixed: Great Tables Row Selection + 5-Min Volume/OI Tracking", Colors.GREEN)
        
        if GREAT_TABLES_AVAILABLE:
            print_colored("✨ Great-Tables: Available for beautiful visualizations", Colors.GREEN)
        elif RICH_AVAILABLE:
            print_colored("✨ Rich: Available for enhanced tables", Colors.GREEN)
        else:
            print_colored("ℹ️  ASCII: Using fallback table formatting", Colors.YELLOW)
        
        print_colored("\n🔧 FINAL FIXES APPLIED:", Colors.CYAN)
        print("  ✅ Great Tables: Fixed row selection using lambda functions")
        print("  ✅ Volume/OI: Using 5-minute timeframe for accurate changes")
        print("  ✅ Tracking: Proper candle-to-candle comparison")
        print("  ✅ Thresholds: Only showing changes >0.1% to reduce noise")
        print("  ✅ Fallbacks: Graceful degradation for missing libraries")
        
        # Display feature overview
        print_colored("\n🚀 KEY FEATURES:", Colors.CYAN)
        print("  🎯 Fixed Great-Tables integration with proper row styling")
        print("  📈 5-minute Volume/OI change tracking with realistic calculations")
        print("  🎨 Rich color-coded tables with change indicators")
        print("  📊 Comprehensive backtesting with detailed statistics")
        print("  🔄 Real-time scanning with precise timing")
        print("  📁 Excel export functionality for analysis")
        
        print_colored("\n📋 USAGE EXAMPLES:", Colors.YELLOW)
        print("  🔴 Live Trading:     python scanner.py")
        print("  🔍 Single Snapshot:  python scanner.py --asof 2025-09-30T14:25")
        print("  📈 Full Day Backtest: python scanner.py --backtest 2025-09-30")
        
        print_colored("="*90, Colors.HEADER)
        
        # Install recommendations
        if not GREAT_TABLES_AVAILABLE:
            print_colored("\n💡 RECOMMENDATION: Install great-tables for best visualization:", Colors.CYAN)
            print("   pip install great-tables")
        
        if not RICH_AVAILABLE:
            print_colored("\n💡 RECOMMENDATION: Install rich for enhanced output:", Colors.CYAN)
            print("   pip install rich")
        
        print_colored("\n🎯 Starting Final Fixed Scanner...", Colors.GREEN)
        
        # Run main function
        main_final_fixed()
        
    except KeyboardInterrupt:
        print_colored("\n\n⚠️  Scanner interrupted by user. Shutting down gracefully...", Colors.YELLOW)
        
        # Show final statistics if available
        if 'scan_count' in globals() and scan_count > 0:
            print_colored(f"📊 Total scans completed: {scan_count}", Colors.CYAN)
        
    except ImportError as e:
        print_colored(f"\n❌ Import Error: {e}", Colors.RED)
        print_colored("💡 Please install required packages:", Colors.YELLOW)
        print("   pip install pandas numpy tqdm rich colorama great-tables")
        print("   pip install truedata-ws openpyxl")
        
    except Exception as e:
        logger.exception(f"❌ Fatal error occurred: {e}")
        print_colored(f"\n💥 Unexpected error: {e}", Colors.RED)
        print_colored("📋 Please check your configuration and try again.", Colors.YELLOW)
        raise
        
    finally:
        # Cleanup and shutdown
        print_colored("\n🔌 Cleaning up resources...", Colors.CYAN)
        
        try:
            # Disconnect TrueData sessions
            if 'tdhist_pool' in globals():
                for i, sess in enumerate(tdhist_pool):
                    try:
                        sess.disconnect()
                    except Exception as cleanup_error:
                        logger.error(f"Error disconnecting session {i}: {cleanup_error}")
            
            print_colored("✅ All TrueData sessions disconnected.", Colors.GREEN)
            
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Final message
        print_colored("\n🎯 Final Fixed Enhanced Option Buyer Scanner v3.2 shutdown complete!", Colors.HEADER)
        print_colored("📊 Thank you for using the Fixed Scanner with Great Tables & 5-Min Tracking!", Colors.GREEN)
        
        # Performance summary if available
        if 'scan_count' in globals() and scan_count > 0:
            uptime_info = f"Completed {scan_count} scans successfully"
            print_colored(f"⏱️  Session Summary: {uptime_info}", Colors.CYAN)

# ========== END OF COMPLETE FIXED SCANNER CODE ==========
